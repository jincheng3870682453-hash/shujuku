# -*- coding: utf-8 -*-
"""
动态数据登记系统 - Flask 后端（全功能版）
端口：5001
"""

import sys

import os
import tempfile
import re
import atexit
import threading
import webbrowser
import hashlib
import secrets
import json
import shutil
import traceback
from collections import Counter
from functools import wraps
from datetime import datetime
from io import BytesIO, StringIO

from flask import Flask, request, jsonify, send_from_directory, g, send_file, session, Response
from flask_cors import CORS

import io


from backend.db_adapter import get_adapter, DatabaseError, reset_adapter
from backend.config import get_db_config, SQLITE_PATH as ADAPTER_SQLITE_PATH, DB_ENGINE

# ──────────────────────── 数据库路径 ────────────────────────
def _resolve_db_path():
    env_path = os.environ.get('REGISTRY_DB_PATH')
    if env_path:
        return env_path
    adapter_path = ADAPTER_SQLITE_PATH
    if adapter_path and adapter_path != "./data/app.db":
        return os.path.abspath(adapter_path)
    preferred = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'registry.db')
    test_dir = os.path.dirname(preferred)
    try:
        _probe = os.path.join(test_dir, '.sqlite_probe')
        with open(_probe, 'w') as f:
            f.write('test')
        os.remove(_probe)
        return preferred
    except Exception:
        fallback = os.path.join(tempfile.gettempdir(), 'registry.db')
        print(f'[INFO] 当前文件系统不支持写入，已回退到: {fallback}')
        return fallback

DATABASE_PATH = _resolve_db_path()
BACKUP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backups', 'db_backups')
os.makedirs(BACKUP_DIR, exist_ok=True)

def _get_real_db_path() -> str:
    """获取实际的数据库文件路径（SQLite 时返回适配层路径，MySQL 返回配置路径）"""
    if DB_ENGINE == 'sqlite':
        adapter_path = os.path.abspath(ADAPTER_SQLITE_PATH)
        if os.path.exists(adapter_path):
            return adapter_path
    return DATABASE_PATH

if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS
else:
    base_path = os.path.dirname(os.path.abspath(__file__))

_templates_path = os.path.join(base_path, 'templates')
_static_path = os.path.join(base_path, 'static')
print(f'[INFO] base_path = {base_path}')
print(f'[INFO] templates 目录: {_templates_path}  存在: {os.path.exists(_templates_path)}')
print(f'[INFO] static 目录: {_static_path}  存在: {os.path.exists(_static_path)}')
if os.path.exists(_templates_path):
    print(f'[INFO] templates 内容: {os.listdir(_templates_path)}')

app = Flask(__name__,
            root_path=base_path,
            template_folder=os.path.join(base_path, 'templates'),
            static_folder=os.path.join(base_path, 'static'))
app.url_map.strict_slashes = False

CORS(app, supports_credentials=True)
app.config['DATABASE'] = DATABASE_PATH
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024
app.secret_key = secrets.token_hex(32)

@app.before_request
def restore_session_from_token():
    """从 Authorization Bearer token 恢复 Flask session，兼容无 Cookie 的 API 客户端（如 Locust 压测）"""
    if session.get('user_id'):
        return  # 已有有效 session，无需恢复

    auth_header = request.headers.get('Authorization', '')
    if not auth_header.startswith('Bearer '):
        return

    token = auth_header[7:].strip()
    if not token.startswith('session-'):
        return

    try:
        user_id = int(token.replace('session-', ''))
    except (ValueError, TypeError):
        return

    # 从数据库恢复用户信息到 session
    try:
        # 注意：before_request 中不能使用 get_db()（它依赖 g 对象，
        # 而 g 在某些 Flask 版本中在 before_request 期间可能未完全初始化）
        from backend.db_adapter import get_adapter as _new_adapter
        adapter = _new_adapter()
        user = adapter.fetch_one("SELECT id, username, role, permissions FROM users WHERE id = ?", (user_id,))
        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            try:
                session['permissions'] = json.dumps(json.loads(user['permissions'] or '[]'))
            except:
                session['permissions'] = '[]'
        adapter.close()
    except Exception as e:
        print(f'[WARN] restore_session_from_token 失败: {e}')

@app.after_request
def add_no_cache_headers(response):
    if request.path.startswith('/static/') or request.path == '/':
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# ──────────────────────── 密码哈希 ────────────────────────
from werkzeug.security import generate_password_hash, check_password_hash

def _legacy_hash_password(password: str) -> str:
    """旧版 SHA-256 无盐哈希（向后兼容已存在的账户）"""
    return hashlib.sha256(password.encode('utf-8')).hexdigest()

def hash_password(password: str) -> str:
    """使用 PBKDF2-SHA256 生成密码哈希"""
    return generate_password_hash(password, method='pbkdf2:sha256')

def verify_password(password: str, stored_hash: str) -> bool:
    """验证密码。自动兼容旧版 SHA-256 和新的 PBKDF2 哈希"""
    if stored_hash.startswith('pbkdf2:') or stored_hash.startswith('scrypt:'):
        return check_password_hash(stored_hash, password)
    # 回退到旧版 SHA-256（向后兼容）
    return _legacy_hash_password(password) == stored_hash

def upgrade_password_if_needed(db, user_id: int, password: str, stored_hash: str):
    """如果用户密码是旧版 SHA-256 哈希，静默升级为 PBKDF2"""
    if not (stored_hash.startswith('pbkdf2:') or stored_hash.startswith('scrypt:')):
        new_hash = hash_password(password)
        db.execute_query("UPDATE users SET password_hash = ? WHERE id = ?", (new_hash, user_id))

# ──────────────────────── 权限常量 ────────────────────────
ALL_PERMISSIONS = [
    'view_data', 'search_data', 'add_data', 'edit_data', 'delete_data',
    'add_field', 'edit_field', 'delete_field', 'batch_add_field',
    'import_excel', 'export_excel',
    'view_logs', 'export_logs', 'clear_logs',
    'audit_center', 'approve_reject',
    'manage_users', 'reset_database',
    'customize_theme', 'view_structure', 'view_stats',
]

ROLE_DEFAULT_PERMISSIONS = {
    'boss': ALL_PERMISSIONS,
    'hr': [
        'view_data', 'search_data', 'add_data', 'edit_data', 'delete_data',
        'add_field', 'edit_field', 'delete_field', 'batch_add_field',
        'import_excel', 'export_excel',
        'view_logs', 'export_logs',
        'audit_center', 'approve_reject',
        'customize_theme', 'view_structure', 'view_stats',
    ],
    'employee': ['view_data', 'search_data', 'view_logs', 'view_structure', 'customize_theme'],
}

def get_user_permissions(user_id):
    db = get_db()
    row = db.fetch_one("SELECT permissions FROM users WHERE id = ?", (user_id,))
    if not row:
        return []
    try:
        perms = json.loads(row['permissions'] or '[]')
        return perms if isinstance(perms, list) else []
    except (json.JSONDecodeError, TypeError):
        return []

def user_has_permission(user_id, perm):
    return perm in get_user_permissions(user_id)

def is_user_audited(user_id):
    db = get_db()
    row = db.fetch_one("SELECT global_audit FROM users WHERE id = ?", (user_id,))
    return bool(row['global_audit']) if row else False

def can_view_all_rows(role):
    """boss 与 hr 属于管理角色，可查看/管理全部数据；员工仅能操作自己创建的数据"""
    return role in ('boss', 'hr')

def require_perm(perm):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            user_id = session.get('user_id')
            if not user_id:
                return jsonify({'error': '请先登录'}), 401
            if not user_has_permission(user_id, perm):
                return jsonify({'error': '权限不足'}), 403
            g.current_user_id = user_id
            g.current_username = session.get('username', '')
            g.current_role = session.get('role', '')
            return f(*args, **kwargs)
        return decorated
    return decorator

def require_role(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            user_id = session.get('user_id')
            if not user_id:
                return jsonify({'error': '请先登录'}), 401
            role = session.get('role', '')
            if role not in roles:
                return jsonify({'error': '权限不足'}), 403
            g.current_user_id = user_id
            g.current_username = session.get('username', '')
            g.current_role = role
            return f(*args, **kwargs)
        return decorated
    return decorator

def require_login(f):
    """仅校验登录状态的装饰器：未登录返回 401，并填充 g.current_* 供日志使用"""
    @wraps(f)
    def decorated(*args, **kwargs):
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': '请先登录'}), 401
        g.current_user_id = user_id
        g.current_username = session.get('username', '')
        g.current_role = session.get('role', '')
        return f(*args, **kwargs)
    return decorated

def log_operation(action, target_type='', target_id='', detail=''):
    try:
        db = get_db()
        db.execute_query(
            "INSERT INTO operations_log (user_id, username, role, action, target_type, target_id, detail) VALUES (?,?,?,?,?,?,?)",
            (g.current_user_id, g.current_username, g.current_role, action, target_type, str(target_id), detail)
        )
    except Exception as e:
        print(f'[WARN] 操作日志写入失败: {e}')

# ──────────────────────── 数据库初始化 ────────────────────────
def _convert_sql_for_engine(sql: str) -> str:
    """将 SQLite 建表语法转换为当前数据库引擎兼容的语法"""
    if DB_ENGINE == 'sqlite':
        return sql
    # MySQL 兼容转换
    converted = sql
    # 1. AUTOINCREMENT → AUTO_INCREMENT
    converted = converted.replace('AUTOINCREMENT', 'AUTO_INCREMENT')
    # 2. INTEGER PRIMARY KEY → INT PRIMARY KEY
    converted = converted.replace('INTEGER PRIMARY KEY AUTO_INCREMENT', 'INT PRIMARY KEY AUTO_INCREMENT')
    # 3. 带约束的 TEXT → VARCHAR（MySQL 不允许 TEXT 有 DEFAULT 或作为索引键）
    converted = _replace_text_for_mysql(converted)
    return converted

def _replace_text_for_mysql(sql: str) -> str:
    """将 TEXT 列替换为 VARCHAR，以兼容 MySQL 的约束限制：
    - MySQL 不允许 TEXT 有 DEFAULT 值（错误 1101）
    - MySQL 不允许 TEXT 作为索引键（错误 1170）
    规则：
    - 带 UNIQUE 或 NOT NULL 的 TEXT → VARCHAR(255)
    - 带 DEFAULT 的 TEXT → VARCHAR(255) + DEFAULT（保留默认值）"""
    import re

    # 带 UNIQUE 或 NOT NULL 的 TEXT → VARCHAR(255)
    pattern_key = r"TEXT\s+((?:(?:NOT\s+NULL|UNIQUE)(?:\s+(?:UNIQUE|NOT\s+NULL))*\s*(?:DEFAULT\s+(?:'[^']*'|\"[^\"]*\"|\([^)]*\)))?\s*)+)"
    sql = re.sub(pattern_key, lambda m: f"VARCHAR(255) {m.group(1).strip()}", sql, flags=re.IGNORECASE)

    # 带 DEFAULT 的 TEXT（无 UNIQUE/NOT NULL）→ VARCHAR(255)，保留 DEFAULT
    # 例如：status TEXT DEFAULT '待审核' → status VARCHAR(255) DEFAULT '待审核'
    pattern_default_only = r"TEXT\s+(DEFAULT\s+(?:'[^']*'|\"[^\"]*\"|\([^)]*\)))"
    sql = re.sub(pattern_default_only, r"VARCHAR(255) \1", sql, flags=re.IGNORECASE)

    return sql

# ── 字段类型映射（MySQL 模式）──
_FIELD_TYPE_MAP = {
    'text': 'VARCHAR(255)', 'number': 'DOUBLE', 'date': 'VARCHAR(255)',
    'select': 'VARCHAR(255)', 'boolean': 'TINYINT(1)', 'textarea': 'TEXT', 'file': 'LONGTEXT'
}


def init_db():
    db = get_adapter()
    ts = db.get_current_timestamp_sql()
    tables = [
        f"CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL, role TEXT NOT NULL DEFAULT 'employee', permissions TEXT DEFAULT '[]', global_audit INTEGER DEFAULT 0, theme TEXT DEFAULT NULL, created_at DATETIME DEFAULT ({ts}))",
        "CREATE TABLE IF NOT EXISTS columns_meta (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL, label TEXT NOT NULL, field_type TEXT NOT NULL DEFAULT 'text', options TEXT, position INTEGER DEFAULT 0)",
        "CREATE TABLE IF NOT EXISTS rows_data (id INTEGER PRIMARY KEY AUTOINCREMENT)",
        f"CREATE TABLE IF NOT EXISTS audit_log (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER, username TEXT, action TEXT, target_table TEXT, target_id TEXT, old_value TEXT, new_value TEXT, created_at DATETIME DEFAULT ({ts}))",
        f"CREATE TABLE IF NOT EXISTS pending_changes (id INTEGER PRIMARY KEY AUTOINCREMENT, row_id INTEGER, column_name TEXT, old_value TEXT, new_value TEXT, change_type TEXT NOT NULL, requested_by INTEGER, status TEXT DEFAULT '待审核', reviewed_by INTEGER, review_comment TEXT, reviewed_at DATETIME, created_at DATETIME DEFAULT ({ts}))",
        f"CREATE TABLE IF NOT EXISTS operations_log (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER, username TEXT, role TEXT, action TEXT, target_type TEXT, target_id TEXT, detail TEXT, created_at DATETIME DEFAULT ({ts}))",
        f"CREATE TABLE IF NOT EXISTS backups (id INTEGER PRIMARY KEY AUTOINCREMENT, filename TEXT UNIQUE, size_bytes INTEGER, created_at DATETIME DEFAULT ({ts}))",
        f"CREATE TABLE IF NOT EXISTS system_config (`key` VARCHAR(255) PRIMARY KEY, `value` TEXT NOT NULL, updated_at DATETIME DEFAULT ({ts}))",
        f"CREATE TABLE IF NOT EXISTS ai_analysis_history (id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL, report TEXT, chat_history TEXT, created_at DATETIME DEFAULT ({ts}), updated_at DATETIME DEFAULT ({ts}))",
    ]
    for t in tables:
        converted = _convert_sql_for_engine(t)
        if converted != t:
            print(f'[DB] SQL converted: {t[:60]}... -> {converted[:60]}...')
        db.execute_query(converted)
    # ── 自动补齐 rows_data 中的系统列 _created_by（用于权限过滤）──
    try:
        existing_sys_cols = db.get_table_columns('rows_data')
        sys_col_names = {r['name'] for r in existing_sys_cols}
    except Exception:
        sys_col_names = set()
    if '_created_by' not in sys_col_names:
        try:
            db.execute_query("ALTER TABLE rows_data ADD COLUMN _created_by VARCHAR(255)")
            print('[INIT] 自动添加系统列: _created_by', flush=True)
        except Exception as e:
            print(f'[INIT] 添加系统列 _created_by 失败: {e}', flush=True)
    # ── 自动补齐 users 表中的 theme 字段（用于按用户存储自定义背景）──
    try:
        user_cols = db.get_table_columns('users')
        user_col_names = {r['name'] for r in user_cols}
    except Exception:
        user_col_names = set()
    if 'theme' not in user_col_names:
        try:
            db.execute_query("ALTER TABLE users ADD COLUMN theme TEXT DEFAULT NULL")
            print('[INIT] 自动添加系统列: users.theme', flush=True)
        except Exception as e:
            print(f'[INIT] 添加系统列 users.theme 失败: {e}', flush=True)
    # ── 自动补齐 rows_data 中缺失的动态列（通过适配层，兼容 SQLite / MySQL）──
    try:
        defined_cols = db.fetch_all("SELECT name, field_type FROM columns_meta ORDER BY position ASC")
        if defined_cols:
            # 通过适配层获取当前表的所有列（SQLite: PRAGMA table_info，MySQL: SHOW COLUMNS）
            try:
                existing_cols = db.get_table_columns('rows_data')
                existing_names = {r['name'] for r in existing_cols}
            except Exception:
                existing_names = set()
            for c in defined_cols:
                cn = c['name']
                if cn not in existing_names:
                    ft = c.get('field_type', 'text')
                    sql_type = _FIELD_TYPE_MAP.get(ft, 'VARCHAR(255)')
                    try:
                        db.execute_query(f"ALTER TABLE rows_data ADD COLUMN {cn} {sql_type}")
                        print(f'[INIT] 自动添加缺失列: {cn} ({sql_type})', flush=True)
                    except Exception as e:
                        print(f'[INIT] 添加列 {cn} 失败: {e}', flush=True)
                else:
                    # 已存在的列，检查是否需要修正类型（file 字段需要 LONGTEXT）
                    ft = c.get('field_type', 'text')
                    expected_type = _FIELD_TYPE_MAP.get(ft, 'VARCHAR(255)')
                    if ft == 'file' and DB_ENGINE != 'sqlite':
                        try:
                            all_cols = db.get_table_columns('rows_data')
                            col_info = [r for r in all_cols if r['name'] == cn]
                            if col_info:
                                current_type = col_info[0]['type']
                                current_type_lower = current_type.lower()
                                if current_type_lower not in ('longtext', 'mediumtext') and (
                                    'varchar' in current_type_lower or ('text' in current_type_lower and 'long' not in current_type_lower)
                                ):
                                    db.execute_query(f"ALTER TABLE rows_data MODIFY COLUMN {cn} {expected_type}")
                                    print(f'[INIT] 修正列类型: {cn} {current_type} → {expected_type}', flush=True)
                        except Exception as e:
                            print(f'[INIT] 检查/修正列 {cn} 类型失败: {e}', flush=True)
    except Exception as e:
        print(f'[INIT] 检查动态列时出错: {e}', flush=True)
    # ── 自动补齐结束 ──

    for username, _, role in [('boss', '123456', 'boss'), ('hr', '123456', 'hr'), ('employee', '123456', 'employee')]:
        if not db.fetch_one("SELECT id FROM users WHERE username = ?", (username,)):
            db.execute_query("INSERT INTO users (username, password_hash, role, permissions) VALUES (?,?,?,?)",
                             (username, hash_password('123456'), role, json.dumps(ROLE_DEFAULT_PERMISSIONS.get(role, []), ensure_ascii=False)))
    for row in db.fetch_all("SELECT id, role, permissions FROM users"):
        default_perms = ROLE_DEFAULT_PERMISSIONS.get(row['role'], [])
        try:
            current = json.loads(row['permissions'] or '[]')
        except:
            current = []
        if not isinstance(current, list):
            current = []
        need = False
        for p in default_perms:
            if p not in current:
                current.append(p)
                need = True
        if need:
            db.execute_query("UPDATE users SET permissions = ? WHERE id = ?", (json.dumps(current, ensure_ascii=False), row['id']))
    # ── 自动修复 pending_changes 表结构 ──
    try:
        cols = db.get_table_columns('pending_changes')
        col_names = {c['name'] for c in cols}
        if 'review_comment' not in col_names:
            try:
                if DB_ENGINE == 'sqlite':
                    db.execute_query("ALTER TABLE pending_changes ADD COLUMN review_comment TEXT")
                else:
                    db.execute_query("ALTER TABLE pending_changes ADD COLUMN review_comment TEXT")
                print('[INIT] pending_changes 添加缺失列 review_comment', flush=True)
            except Exception as e:
                print(f'[INIT] 添加 review_comment 列失败: {e}', flush=True)
        # 修复 status 为 NULL 的旧记录（兼容 MySQL 旧版 DEFAULT 丢失问题）
        db.execute_query("UPDATE pending_changes SET status = '待审核' WHERE status IS NULL OR status = ''")
        print(f'[INIT] pending_changes 表修复完成，现有列: {sorted(col_names)}', flush=True)
    except Exception as e:
        print(f'[INIT] pending_changes 修复失败: {e}', flush=True)
    print('[INFO] 数据库初始化完成')

init_db()

# ──────────────────────── 数据库连接管理 ────────────────────────
def get_db():
    """为当前请求获取独立的数据库连接，存储在 Flask g 对象中"""
    if not hasattr(g, 'db'):
        g.db = get_adapter()
    return g.db

@app.teardown_appcontext
def close_db(exception):
    """请求结束时自动关闭数据库连接"""
    db = g.pop('db', None)
    if db is not None:
        db.close()

# ──────────────────────── 工具函数 ────────────────────────
def get_columns():
    return get_db().fetch_all("SELECT * FROM columns_meta ORDER BY position ASC")

def sqlite_type(field_type):
    return {'text': 'TEXT', 'number': 'REAL', 'date': 'TEXT', 'select': 'TEXT', 'boolean': 'INTEGER', 'textarea': 'TEXT', 'file': 'TEXT'}.get(field_type, 'TEXT')

# ──────────────────────── 共享数据 ────────────────────────
FIELD_TYPE_MAP = _FIELD_TYPE_MAP  # 向后兼容的别名

def convert_field_value(raw, field_type: str):
    """统一的值类型转换，供 api_add_row / api_update_row / api_approve / Excel 导入共用。
    返回 (converted_value, error_message|None)。
    """
    if raw is None or raw == '':
        return None, None
    ft = field_type
    if ft == 'number':
        try:
            return float(raw), None
        except (ValueError, TypeError):
            return None, f'字段值「{raw}」需要是数字'
    elif ft == 'boolean':
        if isinstance(raw, bool):
            return 1 if raw else 0, None
        s = str(raw).strip().lower()
        # 与导入智能识别的布尔词表保持完全一致，避免 "y/对/有" 等被误存为 0
        if s in ('1', 'true', 'yes', 'y', 'on', '是', '对', '正确', '有'):
            return 1, None
        if s in ('0', 'false', 'no', 'n', 'off', '否', '错', '错误', '无'):
            return 0, None
        return 0, None
    elif ft == 'date':
        # 日期时间对象格式化为友好字符串，避免出现 "00:00:00" 尾巴
        if isinstance(raw, datetime):
            if raw.hour == 0 and raw.minute == 0 and raw.second == 0:
                return raw.strftime('%Y-%m-%d'), None
            return raw.strftime('%Y-%m-%d %H:%M:%S'), None
        return str(raw), None
    elif ft == 'file':
        return (json.dumps(raw, ensure_ascii=False) if isinstance(raw, dict) else str(raw)), None
    else:
        # text / select / textarea
        return str(raw), None

def _detect_real_header_row(ws, start_row=1):
    max_row = ws.max_row or 1
    title_kw = ['表', '统计', '汇总', '报表', '信息表', '登记表', '花名册', '人力资源', '员工', '公司', '部门', '年度', '月报', '周报', '日报', '—', '——', '：']
    def _non_empty(row_num):
        return sum(1 for cell in ws[row_num] if cell.value is not None and str(cell.value).strip())
    def _is_title(row_num):
        nn = _non_empty(row_num)
        if nn == 1: return True
        text = ' '.join(str(cell.value).strip() for cell in ws[row_num] if cell.value is not None)
        for kw in title_kw:
            if kw in text: return True
        for ch in text:
            code = ord(ch)
            if (0x1F300 <= code <= 0x1FAFF) or (0x2600 <= code <= 0x27BF): return True
        return False
    checked = min(10, max_row)
    for r in range(start_row, min(start_row + checked, max_row + 1)):
        if _non_empty(r) >= 3 and not _is_title(r):
            return r
    for r in range(start_row, min(start_row + checked, max_row + 1)):
        if _non_empty(r) >= 2: return r
    return start_row

# ──────────────────────── 认证 ────────────────────────
@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json(force=True) or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    if not username or not password:
        return jsonify({'error': '请输入用户名和密码'}), 400
    db = get_db()
    user = db.fetch_one("SELECT id, username, password_hash, role, permissions FROM users WHERE username = ?", (username,))
    if not user or not verify_password(password, user['password_hash']):
        return jsonify({'error': '用户名或密码错误'}), 401
    # 自动升级旧版 SHA-256 密码哈希为 PBKDF2
    upgrade_password_if_needed(db, user['id'], password, user['password_hash'])
    try:
        permissions = json.loads(user['permissions'] or '[]')
    except:
        permissions = []
    session.clear()
    session['user_id'] = user['id']
    session['username'] = user['username']
    session['role'] = user['role']
    session['permissions'] = json.dumps(permissions)
    return jsonify({'token': 'session-' + str(user['id']), 'user': {'id': user['id'], 'username': user['username'], 'role': user['role']}, 'message': '登录成功'})

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'message': '已登出'})

@app.route('/api/health', methods=['GET'])
def api_health():
    return jsonify({'status': 'ok'}), 200

@app.route('/api/me')
def api_me():
    user_id = session.get('user_id')
    if not user_id: return jsonify({'logged_in': False})
    try:
        perms = json.loads(session.get('permissions', '[]'))
    except:
        perms = []
    db = get_db()
    row = db.fetch_one("SELECT COALESCE(global_audit, 0) AS global_audit, theme FROM users WHERE id = ?", (user_id,))
    theme = None
    if row and row.get('theme'):
        try:
            theme = json.loads(row['theme'])
        except Exception:
            theme = None
    return jsonify({'logged_in': True, 'user_id': user_id, 'username': session.get('username', ''), 'role': session.get('role', ''), 'permissions': perms, 'global_audit': bool(row['global_audit']) if row else False, 'theme': theme})

@app.route('/api/theme', methods=['PUT'])
@require_login
def api_save_theme():
    """保存当前登录用户的自定义背景/主题（按用户隔离存储）"""
    data = request.get_json(force=True) or {}
    if not isinstance(data, dict):
        return jsonify({'error': '参数格式错误'}), 400
    # 限制字段范围，避免写入无关数据
    allowed_keys = {'theme', 'texture', 'glass_alpha', 'neon_accent', 'bg_image'}
    payload = {k: data[k] for k in allowed_keys if k in data}
    db = get_db()
    db.execute_query("UPDATE users SET theme = ? WHERE id = ?", (json.dumps(payload, ensure_ascii=False), g.current_user_id))
    log_operation('保存自定义背景', target_type='users', target_id=str(g.current_user_id))
    return jsonify({'message': '已保存', 'theme': payload})

# ──────────────────────── 列管理 ────────────────────────
@app.route('/api/columns', methods=['GET', 'POST'])
def api_columns():
    """GET: 获取所有列定义  POST: 添加新列"""
    if request.method == 'GET':
        if not session.get('user_id'):
            return jsonify({'error': '请先登录'}), 401
        return jsonify(get_columns())

    # === POST ===
    data = request.get_json(force=True) or {}
    # 权限检查
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(user_id, 'add_field'):
        return jsonify({'error': '权限不足'}), 403
    g.current_user_id = user_id
    g.current_username = session.get('username', '')
    g.current_role = session.get('role', '')

    label = (data.get('label') or '').strip()
    field_type = data.get('field_type') or data.get('type', 'text')
    options_raw = data.get('options')
    # options 序列化为 JSON 字符串存入 TEXT 列
    try:
        options = json.dumps(options_raw, ensure_ascii=False) if options_raw is not None else None
    except (TypeError, ValueError) as e:
        print(f'[ERROR] options 序列化失败: {e}, value={repr(options_raw)}')
        return jsonify({'error': '选项数据格式错误'}), 400
    if not label:
        return jsonify({'error': '字段名不能为空'}), 400
    if field_type not in ('text', 'number', 'date', 'select', 'boolean', 'textarea', 'file'):
        return jsonify({'error': f'字段类型不合法: {field_type}'}), 400

    db = get_db()
    role = session.get('role', '')
    if role != 'boss' and is_user_audited(session['user_id']):
        db.execute_query("INSERT INTO pending_changes (row_id, column_name, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?)",
                         (None, '__new_column__', json.dumps(data, ensure_ascii=False), 'add_column', g.current_user_id, '待审核'))
        new_id = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
        log_operation('提交新增字段审核', target_type='pending_changes', target_id=str(new_id), detail=json.dumps(data, ensure_ascii=False))
        return jsonify({'message': '字段添加申请已提交审核', 'pending': True}), 201
    if db.fetch_one("SELECT id FROM columns_meta WHERE label = ?", (label,)):
        return jsonify({'error': f'字段「{label}」已存在'}), 400
    max_pos = (db.fetch_one("SELECT COALESCE(MAX(position), -1) + 1 AS np FROM columns_meta") or {}).get('np', 0)
    col_name = 'col_' + str(max_pos)
    try:
        db.execute_query("INSERT INTO columns_meta (name, label, field_type, options, position) VALUES (?,?,?,?,?)",
                         (col_name, label, field_type, options, max_pos))
    except Exception as e:
        print(f'[ERROR] 添加字段到 columns_meta 失败: {e}')
        return jsonify({'error': f'数据库写入失败: {str(e)}'}), 500
    try:
        db.execute_query(f"ALTER TABLE rows_data ADD COLUMN {col_name} {sqlite_type(field_type)}")
    except Exception as e:
        print(f'[WARN] ALTER TABLE 失败（可能列已存在）: {e}')
    log_operation('添加字段', target_type='columns_meta', target_id=col_name, detail=json.dumps({'label': label}, ensure_ascii=False))
    return jsonify({'name': col_name, 'label': label, 'field_type': field_type}), 201

@app.route('/api/columns/<name>', methods=['DELETE'])
@require_perm('delete_field')
def api_delete_column_by_name(name):
    user_id = session.get('user_id')
    role = session.get('role', '')
    db = get_db()
    col = db.fetch_one("SELECT * FROM columns_meta WHERE name = ?", (name,))
    if not col: return jsonify({'error': '字段不存在'}), 404
    if role != 'boss' and is_user_audited(user_id):
        old_val = json.dumps({'label': col['label'], 'field_type': col['field_type']}, ensure_ascii=False)
        new_val = json.dumps({'label': col['label'], 'action': 'delete'}, ensure_ascii=False)
        db.execute_query("INSERT INTO pending_changes (row_id, column_name, old_value, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?,?)",
                         (None, name, old_val, new_val, 'delete_column', user_id, '待审核'))
        new_id = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
        log_operation('提交删除字段审核', target_type='pending_changes', target_id=str(new_id),
                      detail=json.dumps({'label': col['label']}, ensure_ascii=False))
        return jsonify({'message': '删除字段申请已提交审核', 'pending': True}), 200
    db.execute_query("DELETE FROM columns_meta WHERE name = ?", (name,))
    log_operation('删除字段', target_type='columns_meta', target_id=name, detail=json.dumps({'label': col['label']}, ensure_ascii=False))
    return jsonify({'message': f'字段「{col["label"]}」已删除'})

@app.route('/api/columns/all', methods=['DELETE'])
@require_perm('delete_field')
def api_delete_all_columns():
    user_id = session.get('user_id')
    role = session.get('role', '')
    if role != 'boss': return jsonify({'error': '仅管理员可执行此操作'}), 403
    db = get_db()
    cols = get_columns()
    if is_user_audited(user_id):
        db.execute_query("INSERT INTO pending_changes (row_id, column_name, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?)",
                         (None, '__all__', json.dumps({'count': len(cols)}, ensure_ascii=False), 'clear_columns', user_id, '待审核'))
        new_id = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
        log_operation('提交清空字段审核', target_type='pending_changes', target_id=str(new_id))
        return jsonify({'message': '清空字段申请已提交审核', 'pending': True}), 200
    db.execute_query("DELETE FROM columns_meta")
    db.execute_query("DELETE FROM rows_data")
    log_operation('一键清空所有字段', target_type='columns_meta', detail=json.dumps({'count': len(cols)}, ensure_ascii=False))
    return jsonify({'message': f'已清除 {len(cols)} 个字段及所有数据'})

@app.route('/api/columns/<name>/config', methods=['PUT'])
@require_perm('edit_field')
def api_update_column_config(name):
    """更新字段配置（类型、必填、选项等）"""
    data = request.get_json(force=True) or {}
    user_id = session.get('user_id')
    role = session.get('role', '')
    db = get_db()
    col = db.fetch_one("SELECT * FROM columns_meta WHERE name = ?", (name,))
    if not col:
        return jsonify({'error': '字段不存在'}), 404

    label = data.get('label', col['label'])
    field_type = data.get('type') or data.get('field_type', col['field_type'])
    required = data.get('required', col.get('required', False))
    options_raw = data.get('options')
    if isinstance(required, bool):
        required = 1 if required else 0
    elif isinstance(required, str):
        required = 1 if required.lower() in ('true', '1', 'yes') else 0
    else:
        required = 1 if required else 0

    try:
        options = json.dumps(options_raw, ensure_ascii=False) if options_raw is not None else col.get('options')
    except (TypeError, ValueError) as e:
        print(f'[ERROR] options 序列化失败: {e}')
        return jsonify({'error': '选项数据格式错误'}), 400

    if role != 'boss' and is_user_audited(user_id):
        old_val = json.dumps({'label': col['label'], 'field_type': col['field_type'], 'options': col.get('options')}, ensure_ascii=False)
        audit_data = {'label': label, 'type': field_type, 'field_type': field_type, 'options': options_raw, 'required': required}
        db.execute_query("INSERT INTO pending_changes (row_id, column_name, old_value, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?,?)",
                         (None, name, old_val, json.dumps(audit_data, ensure_ascii=False), 'update_column_config', user_id, '待审核'))
        new_id = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
        log_operation('提交修改字段配置审核', target_type='pending_changes', target_id=str(new_id),
                      detail=json.dumps({'label': label, 'field_type': field_type}, ensure_ascii=False))
        return jsonify({'message': '修改字段配置申请已提交审核', 'pending': True}), 200

    try:
        db.execute_query(
            "UPDATE columns_meta SET label=?, field_type=?, options=? WHERE name=?",
            (label, field_type, options, name)
        )
    except Exception as e:
        print(f'[ERROR] 更新字段配置失败: {e}')
        return jsonify({'error': f'数据库更新失败: {str(e)}'}), 500

    log_operation('更新字段配置', target_type='columns_meta', target_id=name,
                  detail=json.dumps({'label': label, 'field_type': field_type}, ensure_ascii=False))
    return jsonify({'message': '更新成功', 'name': name, 'label': label, 'field_type': field_type})

@app.route('/api/columns/<name>/label', methods=['PUT'])
@require_perm('edit_field')
def api_rename_column_by_name(name):
    data = request.get_json(force=True) or {}
    new_label = (data.get('label') or '').strip()
    if not new_label: return jsonify({'error': '字段名不能为空'}), 400
    user_id = session.get('user_id')
    role = session.get('role', '')
    db = get_db()
    col = db.fetch_one("SELECT * FROM columns_meta WHERE name = ?", (name,))
    if not col: return jsonify({'error': '字段不存在'}), 404
    if role != 'boss' and is_user_audited(user_id):
        old_val = json.dumps({'label': col['label']}, ensure_ascii=False)
        new_val = json.dumps({'label': new_label, 'old_label': col['label']}, ensure_ascii=False)
        db.execute_query("INSERT INTO pending_changes (row_id, column_name, old_value, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?,?)",
                         (None, name, old_val, new_val, 'rename_column', user_id, '待审核'))
        new_id = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
        log_operation('提交重命名字段审核', target_type='pending_changes', target_id=str(new_id),
                      detail=json.dumps({'old': col['label'], 'new': new_label}, ensure_ascii=False))
        return jsonify({'message': '重命名字段申请已提交审核', 'pending': True}), 200
    db.execute_query("UPDATE columns_meta SET label = ? WHERE name = ?", (new_label, name))
    log_operation('重命名字段', target_type='columns_meta', target_id=name, detail=json.dumps({'old': col['label'], 'new': new_label}, ensure_ascii=False))
    return jsonify({'message': f'字段已重命名为「{new_label}」'})

# ──────────────────────── 行数据 ────────────────────────
# ──────────────────────── 文件上传 ────────────────────────
ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'gif', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/api/upload', methods=['POST'])
def api_upload():
    """文件上传接口，保存到 static/uploads/YYYY/MM/DD/ 目录"""
    if not session.get('user_id'):
        return jsonify({'error': '请先登录'}), 401
    file = request.files.get('file')
    if not file or file.filename == '':
        return jsonify({'error': '请选择文件'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': f'不支持此文件格式，请选择: {", ".join(sorted(ALLOWED_EXTENSIONS))}'}), 400

    # 检查文件大小
    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > MAX_FILE_SIZE:
        return jsonify({'error': f'文件大小不能超过 10MB'}), 400

    # 构建保存路径：static/uploads/YYYY/MM/DD/
    today = datetime.now()
    date_path = today.strftime('%Y/%m/%d')
    upload_dir = os.path.join(base_path, 'static', 'uploads', date_path)
    os.makedirs(upload_dir, exist_ok=True)

    # 文件名：时间戳_原文件名（转义特殊字符）
    safe_name = re.sub(r'[^\w\u4e00-\u9fff\-_.]', '_', file.filename.rsplit('.', 1)[0])
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
    timestamp = today.strftime('%Y%m%d_%H%M%S')
    saved_name = f'{timestamp}_{safe_name}.{ext}' if ext else f'{timestamp}_{safe_name}'

    saved_path = os.path.join(upload_dir, saved_name)
    file.save(saved_path)

    # 返回相对于 static 的 URL 路径
    url_path = f'/static/uploads/{date_path}/{saved_name}'
    return jsonify({'url': url_path, 'filename': saved_name, 'size': size}), 200

@app.route('/api/rows', methods=['GET'])
def api_list_rows():
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(session['user_id'], 'view_data'): return jsonify({'error': '权限不足'}), 403
    cols = get_columns()
    if not cols: return jsonify({'data': [], 'total': 0, 'page': 1, 'pageSize': 20})
    page = max(1, request.args.get('page', 1, type=int) or 1)
    page_size = min(100, max(1, request.args.get('pageSize', 20, type=int) or 20))
    col_names = ", ".join(c['name'] for c in cols)
    db = get_db()
    # 具备 view_data 权限的用户均可查看全部数据（增删改仍由权限表控制）
    count_row = db.fetch_one("SELECT COUNT(*) AS cnt FROM rows_data")
    total = count_row['cnt'] if count_row else 0
    offset = (page - 1) * page_size
    rows = db.fetch_all(f"SELECT id, {col_names} FROM rows_data ORDER BY id DESC LIMIT ? OFFSET ?", (page_size, offset))
    return jsonify({
        'data': [{'id': r['id'], **{c['name']: r[c['name']] for c in cols}} for r in rows],
        'total': total,
        'page': page,
        'pageSize': page_size,
    })

@app.route('/api/rows', methods=['POST'])
def api_add_row():
    print("=== POST /api/rows 收到请求 ===", flush=True)
    try:
        if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
        if not user_has_permission(session['user_id'], 'add_data'): return jsonify({'error': '权限不足'}), 403
        data = request.get_json(force=True) or {}
        print(f"[api_add_row] 收到请求体: {json.dumps(data, ensure_ascii=False)}", flush=True)
        cols = get_columns()
        if not cols: return jsonify({'error': '请先添加列再录入数据'}), 400
        role = session.get('role', '')
        g.current_user_id = session['user_id']
        g.current_username = session.get('username', '')
        g.current_role = role
        if role != 'boss' and is_user_audited(session['user_id']):
            db = get_db()
            db.execute_query("INSERT INTO pending_changes (row_id, column_name, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?)",
                             (None, '__full_row__', json.dumps(data, ensure_ascii=False), 'insert', g.current_user_id, '待审核'))
            return jsonify({'message': '数据已提交审核', 'pending': True}), 201
        db = get_db()
        values, col_names = [], []
        for c in cols:
            raw = data.get(c['name'])
            col_names.append(c['name'])
            val, err = convert_field_value(raw, c['field_type'])
            values.append(val)
        print(f"[api_add_row] 列名: {col_names}", flush=True)
        print(f"[api_add_row] 值: {values}", flush=True)
        
        # 自动补齐 rows_data 中缺失的动态列（通过适配层，兼容 SQLite / MySQL）
        try:
            existing_cols = db.get_table_columns('rows_data')
            existing_names = {r['name'] for r in existing_cols}
        except Exception:
            existing_names = set()
        for c in cols:
            cn = c['name']
            if cn not in existing_names:
                ft = c.get('field_type', 'text')
                sql_type = FIELD_TYPE_MAP.get(ft, 'VARCHAR(255)')
                try:
                    db.execute_query(f"ALTER TABLE rows_data ADD COLUMN {cn} {sql_type}")
                    print(f"[api_add_row] 自动添加缺失列: {cn} ({sql_type})", flush=True)
                except Exception as e:
                    print(f"[api_add_row] 添加列 {cn} 失败: {e}", flush=True)
        
        sql = f"INSERT INTO rows_data ({', '.join(col_names)}, _created_by) VALUES ({', '.join('?' for _ in col_names)}, ?)"
        print(f"[api_add_row] SQL: {sql}", flush=True)
        created_by = session.get('username', '')
        db.execute_query(sql, tuple(values) + (created_by,))
        new_id = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
        log_operation('新增数据', target_type='rows_data', target_id=str(new_id))
        print(f"[api_add_row] 新增成功，id={new_id}", flush=True)
        return jsonify({'id': new_id, 'message': '新增成功'}), 201
    except Exception as e:
        print(f"[api_add_row] 异常: {e}", flush=True)
        traceback.print_exc()
        return jsonify({'error': f'服务器内部错误: {str(e)}'}), 500

@app.route('/api/rows/<int:row_id>', methods=['PUT'])
def api_update_row(row_id):
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(session['user_id'], 'edit_data'): return jsonify({'error': '权限不足'}), 403
    data = request.get_json(force=True) or {}
    print(f'[api_update_row] row_id={row_id}, data={json.dumps(data, ensure_ascii=False)}')
    cols = get_columns()
    db = get_db()
    # 权限检查：非管理角色只能修改自己创建的数据
    role = session.get('role', '')
    username = session.get('username', '')
    if not can_view_all_rows(role):
        existing = db.fetch_one("SELECT id FROM rows_data WHERE id = ? AND _created_by = ?", (row_id, username))
    else:
        existing = db.fetch_one("SELECT id FROM rows_data WHERE id = ?", (row_id,))
    if not existing:
        print(f'[api_update_row] record {row_id} not found')
        return jsonify({'error': f'记录 {row_id} 不存在或无权修改'}), 404
    g.current_user_id = session['user_id']
    g.current_username = session.get('username', '')
    g.current_role = role
    if role != 'boss' and is_user_audited(session['user_id']):
        print(f'[api_update_row] audit enabled, inserting to pending_changes')
        for c in cols:
            if c['name'] not in data: continue
            db.execute_query("INSERT INTO pending_changes (row_id, column_name, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?)",
                             (row_id, c['name'], json.dumps(data[c['name']], ensure_ascii=False), 'update', g.current_user_id, '待审核'))
        return jsonify({'message': '修改已提交审核', 'pending': True})
    set_parts, values = [], []
    for c in cols:
        if c['name'] not in data: continue
        val, err = convert_field_value(data[c['name']], c['field_type'])
        if err and c['field_type'] == 'number':
            return jsonify({'error': f'字段「{c["label"]}」需要数字'}), 400
        set_parts.append(f"{c['name']} = ?")
        values.append(val)
    if not set_parts:
        print(f'[api_update_row] no matching columns found in data. cols={[c["name"] for c in cols]}, data keys={list(data.keys())}')
        return jsonify({'error': '没有可更新的字段'}), 400
    values.append(row_id)
    sql = f"UPDATE rows_data SET {', '.join(set_parts)} WHERE id = ?"
    print(f'[api_update_row] SQL: {sql}, params: {values}')
    try:
        db.execute_query(sql, tuple(values))
        print(f'[api_update_row] update success')
    except Exception as e:
        print(f'[api_update_row] update failed: {e}')
        return jsonify({'error': f'更新失败: {str(e)}'}), 500
    log_operation('修改数据', target_type='rows_data', target_id=str(row_id))
    return jsonify({'message': '编辑成功'})

@app.route('/api/rows/all', methods=['DELETE'])
@require_perm('delete_data')
def api_delete_all_rows():
    """一键清空所有数据（仅清空 rows_data，保留字段结构）"""
    user_id = session.get('user_id')
    role = session.get('role', '')
    if role != 'boss':
        return jsonify({'error': '仅管理员可执行此操作'}), 403
    db = get_db()
    count_row = db.fetch_one("SELECT COUNT(*) AS cnt FROM rows_data")
    count = count_row['cnt'] if count_row else 0
    if is_user_audited(user_id):
        db.execute_query("INSERT INTO pending_changes (row_id, column_name, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?)",
                         (None, '__all_rows__', json.dumps({'count': count}, ensure_ascii=False), 'clear_rows', user_id, '待审核'))
        new_id = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
        log_operation('提交清空数据审核', target_type='pending_changes', target_id=str(new_id))
        return jsonify({'message': '清空数据申请已提交审核', 'pending': True}), 200
    db.execute_query("DELETE FROM rows_data")
    log_operation('一键清空所有数据', target_type='rows_data', detail=json.dumps({'count': count}, ensure_ascii=False))
    return jsonify({'message': f'已清空所有数据（共 {count} 条）', 'count': count})

@app.route('/api/rows/<int:row_id>', methods=['DELETE'])
def api_delete_row(row_id):
    user_id = session.get('user_id')
    if not user_id: return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(user_id, 'delete_data'): return jsonify({'error': '权限不足'}), 403
    db = get_db()
    role = session.get('role', '')
    username = session.get('username', '')
    # 权限检查：非管理角色只能删除自己创建的数据
    if not can_view_all_rows(role):
        existing_row = db.fetch_one("SELECT * FROM rows_data WHERE id = ? AND _created_by = ?", (row_id, username))
    else:
        existing_row = db.fetch_one("SELECT * FROM rows_data WHERE id = ?", (row_id,))
    if not existing_row: return jsonify({'error': '记录不存在或无权删除'}), 404
    g.current_user_id = user_id; g.current_username = session.get('username', ''); g.current_role = role
    # 审核流：HR 角色始终提交审核；员工按全局审核开关决定；boss 直接执行
    if role != 'boss' and (role == 'hr' or is_user_audited(user_id)):
        row_snapshot = json.dumps(dict(existing_row), ensure_ascii=False, default=str)
        db.execute_query("INSERT INTO pending_changes (row_id, column_name, old_value, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?,?)",
                         (row_id, '__delete_row__', row_snapshot, '{}', 'delete', user_id, '待审核'))
        new_id = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
        log_operation('提交删除数据审核', target_type='pending_changes', target_id=str(new_id))
        return jsonify({'message': '删除申请已提交审核', 'pending': True}), 200
    db.execute_query("DELETE FROM rows_data WHERE id = ?", (row_id,))
    log_operation('删除数据', target_type='rows_data', target_id=str(row_id))
    return jsonify({'message': '已删除'})

@app.route('/api/rows/batch-delete', methods=['POST'])
def api_batch_delete_rows():
    user_id = session.get('user_id')
    if not user_id: return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(user_id, 'delete_data'): return jsonify({'error': '权限不足'}), 403
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    if not ids or not isinstance(ids, list):
        return jsonify({'error': '请提供要删除的ID列表'}), 400
    db = get_db()
    role = session.get('role', '')
    username = session.get('username', '')
    g.current_user_id = user_id; g.current_username = username; g.current_role = role
    deleted = 0
    for row_id in ids:
        try:
            if not can_view_all_rows(role):
                existing = db.fetch_one("SELECT * FROM rows_data WHERE id = ? AND _created_by = ?", (row_id, username))
            else:
                existing = db.fetch_one("SELECT * FROM rows_data WHERE id = ?", (row_id,))
            if not existing:
                continue
            # 审核流：HR 角色始终提交审核；员工按全局审核开关决定；boss 直接执行
            if role != 'boss' and (role == 'hr' or is_user_audited(user_id)):
                row_snapshot = json.dumps(dict(existing), ensure_ascii=False, default=str)
                db.execute_query("INSERT INTO pending_changes (row_id, column_name, old_value, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?,?)",
                                 (row_id, '__delete_row__', row_snapshot, '{}', 'delete', user_id, '待审核'))
                new_pid = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
                log_operation('批量提交删除数据审核', target_type='pending_changes', target_id=str(new_pid))
            else:
                db.execute_query("DELETE FROM rows_data WHERE id = ?", (row_id,))
                log_operation('批量删除数据', target_type='rows_data', target_id=str(row_id))
            deleted += 1
        except Exception:
            continue
    return jsonify({'message': f'成功删除 {deleted} 条记录'})

# ──────────────────────── 审核 ────────────────────────
@app.route('/api/audit', methods=['GET'])
def api_audit_list():
    try:
        if not session.get('user_id'):
            return jsonify({'error': '请先登录'}), 401
        db = get_db()
        status = request.args.get('status')
        page = max(1, request.args.get('page', 1, type=int) or 1)
        page_size = min(100, max(1, request.args.get('pageSize', 20, type=int) or 20))
        role = session.get('role', '')
        base_sql = "SELECT pc.id, pc.row_id, pc.column_name, pc.old_value, pc.new_value, pc.change_type, pc.requested_by, pc.status, pc.reviewed_by, pc.review_comment, pc.reviewed_at, pc.created_at, u.username AS applicant, u.role AS applicant_role FROM pending_changes pc LEFT JOIN users u ON pc.requested_by = u.id"
        conditions = []
        params = []
        if role != 'boss':
            conditions.append("(pc.requested_by = ? OR pc.reviewed_by = ?)")
            params.extend([session['user_id'], session['user_id']])
        if status:
            status_map = {'pending': '待审核', 'approved': '已通过', 'rejected': '已驳回'}
            conditions.append("pc.status = ?")
            params.append(status_map.get(status, status))
        if conditions:
            base_sql += " WHERE " + " AND ".join(conditions)
        count_sql = f"SELECT COUNT(*) AS cnt FROM ({base_sql}) AS _sub"
        total = (db.fetch_one(count_sql, tuple(params)) or {}).get('cnt', 0)
        rows = db.fetch_all(base_sql + " ORDER BY pc.created_at DESC LIMIT ? OFFSET ?", tuple(params) + (page_size, (page - 1) * page_size))
        result_rows = []
        for r in (rows or []):
            d = dict(r)
            d['type'] = d.get('change_type', '')
            # 过滤审核记录中的密码明文（new_value 和 old_value）
            _PASSWORD_KEYS = {'password', 'password_hash', 'new_password', 'old_password', 'confirm_password'}
            for _field in ('new_value', 'old_value'):
                if d.get(_field):
                    try:
                        _v = json.loads(d[_field])
                        if isinstance(_v, dict):
                            _redacted = False
                            for _pk in _PASSWORD_KEYS:
                                if _pk in _v:
                                    _v[_pk] = '******'
                                    _redacted = True
                            if _redacted:
                                d[_field] = json.dumps(_v, ensure_ascii=False)
                    except (json.JSONDecodeError, TypeError):
                        pass
            result_rows.append(d)
        return jsonify({'data': result_rows, 'total': total, 'page': page, 'pageSize': page_size})
    except Exception as e:
        print(f'[ERROR] GET /api/audit 异常: {e}', flush=True)
        traceback.print_exc()
        return jsonify({'error': f'服务器内部错误: {str(e)}'}), 500

@app.route('/api/audit/count', methods=['GET'])
def api_audit_count():
    user_id = session.get('user_id')
    if not user_id: return jsonify({'error': '请先登录'}), 401
    db = get_db()
    role = session.get('role', '')
    row = db.fetch_one("SELECT COUNT(*) AS pending FROM pending_changes WHERE status = '待审核'" + ('' if role == 'boss' else ' AND requested_by = ?'), (() if role == 'boss' else (user_id,)))
    return jsonify({'pending': row['pending'] if row else 0})

@app.route('/api/audit/<int:change_id>/approve', methods=['POST'])
@require_perm('approve_reject')
def api_approve(change_id):
    db = get_db()
    pc = db.fetch_one("SELECT * FROM pending_changes WHERE id = ? AND status = '待审核'", (change_id,))
    if not pc: return jsonify({'error': '审核记录不存在或已处理'}), 404
    ct = pc['change_type'] or ''
    row_id = pc['row_id']
    new_value_str = pc['new_value'] or '{}'
    try:
        new_value = json.loads(new_value_str) if isinstance(new_value_str, str) else new_value_str
    except (json.JSONDecodeError, TypeError):
        new_value = {}
    
    # ── 根据 change_type 执行实际操作 ──
    try:
        if ct == 'delete':
            # 删除数据行
            if row_id:
                db.execute_query("DELETE FROM rows_data WHERE id = ?", (row_id,))
                log_operation('审核后删除数据', target_type='rows_data', target_id=str(row_id))
        elif ct == 'delete_column':
            col_name = pc['column_name'] or ''
            if col_name:
                db.execute_query("DELETE FROM columns_meta WHERE name = ?", (col_name,))
                log_operation('审核后删除字段', target_type='columns_meta', target_id=col_name)
        elif ct in ('update_column', 'update_column_config'):
            col_name = pc['column_name'] or ''
            if col_name and isinstance(new_value, dict):
                label = new_value.get('label')
                field_type = new_value.get('type') or new_value.get('field_type')
                options = new_value.get('options')
                if label and field_type:
                    options_json = json.dumps(options, ensure_ascii=False) if options is not None else None
                    db.execute_query("UPDATE columns_meta SET label=?, field_type=?, options=? WHERE name=?",
                                     (label, field_type, options_json, col_name))
                log_operation('审核后修改字段配置', target_type='columns_meta', target_id=col_name)
        elif ct == 'rename_column':
            col_name = pc['column_name'] or ''
            if col_name and isinstance(new_value, dict):
                new_label = new_value.get('label', '')
                if new_label:
                    db.execute_query("UPDATE columns_meta SET label = ? WHERE name = ?", (new_label, col_name))
                log_operation('审核后重命名字段', target_type='columns_meta', target_id=col_name)
        elif ct == 'clear_columns':
            db.execute_query("DELETE FROM columns_meta")
            db.execute_query("DELETE FROM rows_data")
            log_operation('审核后清空所有字段')
        elif ct == 'clear_rows':
            db.execute_query("DELETE FROM rows_data")
            log_operation('审核后清空所有数据')
        elif ct == 'clear_logs':
            db.execute_query("DELETE FROM operations_log")
            log_operation('审核后清空日志')
        elif ct == 'add_user':
            if isinstance(new_value, dict):
                username = new_value.get('username', '')
                password = new_value.get('password', '')
                role = new_value.get('role', 'employee')
                if username and password and len(password) >= 6:
                    if not db.fetch_one("SELECT id FROM users WHERE username = ?", (username,)):
                        perms = json.dumps(ROLE_DEFAULT_PERMISSIONS.get(role, []), ensure_ascii=False)
                        db.execute_query("INSERT INTO users (username, password_hash, role, permissions) VALUES (?,?,?,?)",
                                         (username, hash_password(password), role, perms))
                        log_operation('审核后添加用户', target_type='users', detail=json.dumps({'username': username, 'role': role}, ensure_ascii=False))
        elif ct == 'update_user':
            if isinstance(new_value, dict):
                uid = new_value.get('id') or row_id
                username = new_value.get('username', '')
                role = new_value.get('role', 'employee')
                password = new_value.get('password')
                if uid and username:
                    if password:
                        db.execute_query("UPDATE users SET username=?, password_hash=?, role=? WHERE id=?", (username, hash_password(password), role, uid))
                    else:
                        db.execute_query("UPDATE users SET username=?, role=? WHERE id=?", (username, role, uid))
                    log_operation('审核后修改用户', target_type='users', target_id=str(uid))
        elif ct == 'delete_user':
            uid = row_id or (new_value.get('id') if isinstance(new_value, dict) else None)
            if uid:
                u = db.fetch_one("SELECT username FROM users WHERE id=? AND role!='boss'", (uid,))
                if u:
                    db.execute_query("DELETE FROM users WHERE id=?", (uid,))
                    log_operation('审核后删除用户', target_type='users', target_id=str(uid))
        elif ct == 'add_column':
            if isinstance(new_value, dict):
                label = (new_value.get('label') or '').strip()
                field_type = new_value.get('field_type') or new_value.get('type', 'text')
                if label and not db.fetch_one("SELECT id FROM columns_meta WHERE label = ?", (label,)):
                    max_pos = (db.fetch_one("SELECT COALESCE(MAX(position), -1) + 1 AS np FROM columns_meta") or {}).get('np', 0)
                    col_name = 'col_' + str(max_pos)
                    try:
                        options_raw = new_value.get('options')
                        options_json = json.dumps(options_raw, ensure_ascii=False) if options_raw is not None else None
                        db.execute_query("INSERT INTO columns_meta (name, label, field_type, options, position) VALUES (?,?,?,?,?)",
                                         (col_name, label, field_type, options_json, max_pos))
                        db.execute_query(f"ALTER TABLE rows_data ADD COLUMN {col_name} {sqlite_type(field_type)}")
                        log_operation('审核后添加字段', target_type='columns_meta', target_id=col_name)
                    except Exception as e:
                        print(f'[APPROVE] add_column 失败: {e}')
        elif ct == 'insert':
            # 新增数据行 — 参考 api_add_row 逻辑
            if isinstance(new_value, dict):
                cols = get_columns()
                values, col_names = [], []
                for c in cols:
                    col_names.append(c['name'])
                    val, _ = convert_field_value(new_value.get(c['name']), c['field_type'])
                    values.append(val)
                try:
                    existing_cols = db.get_table_columns('rows_data')
                    existing_names = {r['name'] for r in existing_cols}
                except: existing_names = set()
                for c in cols:
                    cn = c['name']
                    if cn not in existing_names:
                        ft = c.get('field_type', 'text')
                        sql_type = FIELD_TYPE_MAP.get(ft, 'VARCHAR(255)')
                        try: db.execute_query(f"ALTER TABLE rows_data ADD COLUMN {cn} {sql_type}")
                        except: pass
                # 获取请求用户的用户名作为 _created_by
                req_user = db.fetch_one("SELECT username FROM users WHERE id = ?", (int(pc['requested_by']),)) if pc.get('requested_by') else None
                created_by_user = req_user['username'] if req_user else ''
                sql = f"INSERT INTO rows_data ({', '.join(col_names)}, _created_by) VALUES ({', '.join('?' for _ in col_names)}, ?)"
                db.execute_query(sql, tuple(values) + (created_by_user,))
                new_id = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
                log_operation('审核后新增数据', target_type='rows_data', target_id=str(new_id))
        elif ct == 'update':
            # 更新数据行 — 参考 api_update_row 逻辑
            if row_id and isinstance(new_value, dict):
                cols = get_columns()
                set_parts, update_values = [], []
                for c in cols:
                    if c['name'] not in new_value: continue
                    val, _ = convert_field_value(new_value[c['name']], c['field_type'])
                    set_parts.append(f"{c['name']} = ?")
                    update_values.append(val)
                if set_parts:
                    update_values.append(row_id)
                    db.execute_query(f"UPDATE rows_data SET {', '.join(set_parts)} WHERE id = ?", tuple(update_values))
                    log_operation('审核后修改数据', target_type='rows_data', target_id=str(row_id))
        # ── reset_database / restore_backup / import 等复杂操作暂不自动执行，仅标记通过 ──
        else:
            print(f'[APPROVE] 未知 change_type: {ct}，仅标记通过')
    except Exception as e:
        print(f'[APPROVE] 执行操作失败 ({ct}): {e}')
        traceback.print_exc()
    
    db.execute_query("UPDATE pending_changes SET status = '已通过', reviewed_by = ?, reviewed_at = datetime('now') WHERE id = ?", (g.current_user_id, change_id))
    log_operation('审核通过', target_type='pending_changes', target_id=str(change_id))
    return jsonify({'message': '已通过'})

@app.route('/api/audit/<int:change_id>/reject', methods=['POST'])
@require_perm('approve_reject')
def api_reject(change_id):
    data = request.get_json(force=True) or {}
    db = get_db()
    pc = db.fetch_one("SELECT * FROM pending_changes WHERE id = ? AND status = '待审核'", (change_id,))
    if not pc: return jsonify({'error': '审核记录不存在或已处理'}), 404
    db.execute_query("UPDATE pending_changes SET status = '已驳回', reviewed_by = ?, review_comment = ?, reviewed_at = datetime('now') WHERE id = ?", (g.current_user_id, data.get('comment', ''), change_id))
    log_operation('审核驳回', target_type='pending_changes', target_id=str(change_id))
    return jsonify({'message': '已驳回'})

# ──────────────────────── 日志 ────────────────────────
@app.route('/api/logs', methods=['GET'])
def api_list_logs():
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(session['user_id'], 'view_logs'): return jsonify({'error': '权限不足'}), 403
    db = get_db()
    page = max(1, request.args.get('page', 1, type=int) or 1)
    page_size = min(100, max(1, request.args.get('pageSize', 20, type=int) or 20))
    keyword = request.args.get('keyword', '').strip()
    conditions, params = [], []
    if keyword: conditions.append("(username LIKE ? OR action LIKE ? OR detail LIKE ?)"); kw = f'%{keyword}%'; params.extend([kw, kw, kw])
    base = "SELECT * FROM operations_log" + ((" WHERE " + " AND ".join(conditions)) if conditions else "")
    total = (db.fetch_one(f"SELECT COUNT(*) AS cnt FROM ({base}) AS t", tuple(params)) or {}).get('cnt', 0)
    rows = db.fetch_all(base + " ORDER BY created_at DESC LIMIT ? OFFSET ?", tuple(params) + (page_size, (page - 1) * page_size))
    return jsonify({'data': [dict(r) for r in rows], 'total': total, 'page': page, 'pageSize': page_size})

@app.route('/api/logs/all', methods=['DELETE'])
@require_perm('clear_logs')
def api_clear_all_logs():
    user_id = session.get('user_id')
    role = session.get('role', '')
    db = get_db()
    if role != 'boss' and is_user_audited(user_id):
        db.execute_query("INSERT INTO pending_changes (row_id, column_name, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?)",
                         (None, '__clear_logs__', '{}', 'clear_logs', user_id, '待审核'))
        new_id = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
        log_operation('提交清空日志审核', target_type='pending_changes', target_id=str(new_id))
        return jsonify({'message': '清空日志申请已提交审核', 'pending': True}), 200
    db.execute_query("DELETE FROM operations_log")
    log_operation('清空所有操作日志')
    return jsonify({'message': '所有日志已清空'})

@app.route('/api/logs/export', methods=['GET'])
def api_export_logs():
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(session['user_id'], 'export_logs'): return jsonify({'error': '权限不足'}), 403
    db = get_db()
    rows = db.fetch_all("SELECT * FROM operations_log ORDER BY created_at DESC")
    output = StringIO()
    output.write("操作时间\t用户\t角色\t操作\t目标类型\t详情\n")
    for r in rows: output.write(f"{r['created_at']}\t{r['username']}\t{r['role']}\t{r['action']}\t{r['target_type']}\t{r['detail'] or ''}\n")
    content = output.getvalue(); output.close()
    return Response(content.encode('utf-8-sig'), mimetype='text/plain; charset=utf-8',
                    headers={'Content-Disposition': f'attachment; filename=operation_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'})

# ──────────────────────── 统计 ────────────────────────
@app.route('/api/stats', methods=['GET'])
def api_stats():
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(session['user_id'], 'view_stats'): return jsonify({'error': '权限不足'}), 403
    g.current_user_id = session['user_id']; g.current_username = session.get('username', ''); g.current_role = session.get('role', '')
    field_key = (request.args.get('field') or '').strip()
    if not field_key: return jsonify({'error': '请指定统计字段'}), 400
    db = get_db()
    col = db.fetch_one("SELECT * FROM columns_meta WHERE name = ?", (field_key,))
    if not col: return jsonify({'error': '字段不存在'}), 404
    # 权限过滤：非管理角色只能看到自己创建的数据
    role = session.get('role', '')
    if not can_view_all_rows(role):
        username = session.get('username', '')
        rows = db.fetch_all(f"SELECT {field_key} FROM rows_data WHERE {field_key} IS NOT NULL AND {field_key} != '' AND _created_by = ?", (username,))
    else:
        rows = db.fetch_all(f"SELECT {field_key} FROM rows_data WHERE {field_key} IS NOT NULL AND {field_key} != ''")
    values = [r[field_key] for r in rows]
    if not values: return jsonify({'field_key': field_key, 'field_label': col['label'], 'total': 0, 'items': []})
    counter = Counter()
    for v in values: counter[str(v)] += 1
    items = [{'name': k, 'value': v} for k, v in counter.most_common()]
    return jsonify({'field_key': field_key, 'field_label': col['label'], 'total': sum(counter.values()), 'items': items})

@app.route('/api/stats/fields', methods=['GET'])
def api_stats_fields():
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(session['user_id'], 'view_stats'): return jsonify({'error': '权限不足'}), 403
    cols = get_columns()
    return jsonify([{'key': c['name'], 'label': c['label'], 'type': c['field_type']} for c in cols])

# ──────────────────────── 导出/导入 Excel ────────────────────────
@app.route('/api/export', methods=['GET'])
def api_export():
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(session['user_id'], 'export_excel'): return jsonify({'error': '权限不足'}), 403
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': '未安装 openpyxl'}), 500
    cols = get_columns()
    col_names = ", ".join(c['name'] for c in cols) if cols else '*'
    db = get_db()
    role = session.get('role', '')
    if not can_view_all_rows(role):
        username = session.get('username', '')
        rows = db.fetch_all(f"SELECT id, {col_names} FROM rows_data WHERE _created_by = ? ORDER BY id ASC", (username,))
    else:
        rows = db.fetch_all(f"SELECT id, {col_names} FROM rows_data ORDER BY id ASC")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '数据登记表'
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    headers = ['ID'] + [c['label'] for c in cols]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h); c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal='center')
    for ri, row in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=row['id'])
        for ci, col in enumerate(cols, 2): ws.cell(row=ri, column=ci, value=row[col['name']])
    output = BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'数据登记表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

@app.route('/api/import', methods=['POST'])
def api_import():
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(session['user_id'], 'import_excel'): return jsonify({'error': '权限不足'}), 403
    g.current_user_id = session['user_id']; g.current_username = session.get('username', ''); g.current_role = session.get('role', '')
    try: import openpyxl
    except ImportError: return jsonify({'error': '未安装 openpyxl'}), 500
    file = request.files.get('file')
    if not file: return jsonify({'error': '请上传文件'}), 400
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    user_header = request.form.get('header_row', type=int, default=1) or 1
    if user_header == 1:
        header_row = _detect_real_header_row(ws, 1)
    else:
        header_row = user_header
    data_start = header_row + 1
    rows_data = [row for row in ws.iter_rows(min_row=data_start, values_only=True) if any(cell is not None for cell in row)]
    if not rows_data: return jsonify({'message': '文件中没有数据'}), 200
    db = get_db()
    excel_headers = []
    header_row_data = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    if header_row_data and header_row_data[0]:
        excel_headers = [str(c).strip() if c is not None else '' for c in header_row_data[0]]
    else:
        excel_headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[header_row]]
    def _detect_column_type(values):
        """智能识别字段类型：分析该列全部数据后决定类型。

        识别顺序：布尔词 → 数字 → 日期 → 下拉选项 → 文本。
        返回 (field_type, options)。
        """
        non_empty = [v for v in values if v is not None and str(v).strip() != '']
        if not non_empty:
            return 'text', None
        def _is_boolean_word(v):
            if isinstance(v, bool): return True
            return str(v).strip().lower() in ('是', '否', 'true', 'false', 'yes', 'no', 'y', 'n', '对', '错', '正确', '错误', '有', '无')
        def _is_number(v):
            if isinstance(v, bool): return False
            if isinstance(v, (int, float)): return True
            s = str(v).strip()
            if not s: return False
            # 手机号 / 身份证号等纯数字长串视为文本，避免误判为数字
            if s.isdigit() and len(s) >= 10: return False
            try:
                float(s)
                return True
            except ValueError:
                return False
        def _is_date(v):
            if isinstance(v, datetime): return True
            if isinstance(v, (int, float, bool)): return False
            s = str(v).strip()
            if not s: return False
            for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日',
                        '%m/%d/%Y', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S',
                        '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M'):
                try:
                    datetime.strptime(s, fmt)
                    return True
                except ValueError:
                    continue
            return False
        # 1) 全部是布尔词 → 布尔/开关
        if all(_is_boolean_word(v) for v in non_empty):
            return 'boolean', None
        # 2) 全部可转数字 → 数字
        if all(_is_number(v) for v in non_empty):
            return 'number', None
        # 3) 全部是日期 → 日期
        if all(_is_date(v) for v in non_empty):
            return 'date', None
        # 4) 取值有限且重复率高 → 下拉选择（自动生成选项）
        distinct, seen = [], set()
        for v in non_empty:
            s = str(v).strip()
            if s not in seen:
                seen.add(s)
                distinct.append(s)
        if 1 < len(distinct) <= 20 and len(distinct) <= len(non_empty) * 0.6:
            options = json.dumps([{'label': s, 'value': s} for s in distinct], ensure_ascii=False)
            return 'select', options
        return 'text', None

    def auto_create_columns(headers, rows_data):
        new_names = []
        used_names = set()
        for idx, h in enumerate(headers):
            if not h: continue
            if idx == 0 and h.lower() == 'id': continue
            name = re.sub(r'[^\w\u4e00-\u9fff]', '_', h)
            name = re.sub(r'_+', '_', name).strip('_').lower()
            if not name or name[0].isdigit(): name = 'col_' + name
            if db.fetch_one("SELECT id FROM columns_meta WHERE name = ? OR label = ?", (name, h)): continue
            # 边界情况：同一 Excel 内出现重名表头（如两个「备注」）时追加序号，
            # 避免命中 columns_meta.name 的 UNIQUE 约束导致导入崩溃
            base, n = name, 2
            while name in used_names:
                name = f"{base}_{n}"
                n += 1
            used_names.add(name)
            max_pos = (db.fetch_one("SELECT COALESCE(MAX(position), -1) + 1 AS np FROM columns_meta") or {}).get('np', 0)
            # 智能识别字段类型：分析该列全部数据
            values = [row[idx] if idx < len(row) else None for row in rows_data]
            field_type, options = _detect_column_type(values)
            db.execute_query("INSERT INTO columns_meta (name, label, field_type, options, position) VALUES (?,?,?,?,?)", (name, h, field_type, options, max_pos + idx))
            try: db.execute_query(f"ALTER TABLE rows_data ADD COLUMN {name} {sqlite_type(field_type)}")
            except: pass
            new_names.append(name)
        return new_names, get_columns()
    cols = get_columns()
    new_fields = []
    if not cols:
        new_fields, cols = auto_create_columns(excel_headers, rows_data)
    if not cols: return jsonify({'error': '无法创建字段定义'}), 400
    col_names = [c['name'] for c in cols]
    places = ", ".join("?" for _ in col_names)
    col_list = ", ".join(col_names)
    created_by = session.get('username', '')
    ok, fail = 0, 0
    for row in rows_data:
        vals = []
        for i, c in enumerate(cols):
            v = row[i] if i < len(row) else None
            val, _ = convert_field_value(v, c['field_type'])
            vals.append(val)
        try: db.execute_query(f"INSERT INTO rows_data ({col_list}, _created_by) VALUES ({places}, ?)", tuple(vals) + (created_by,)); ok += 1
        except: fail += 1
    log_operation('导入Excel', detail=json.dumps({'import_count': ok, 'fail_count': fail, 'new_fields': len(new_fields)}, ensure_ascii=False))
    msg = f'成功导入 {ok} 条数据'
    if fail > 0: msg += f'，失败 {fail} 条'
    if new_fields: msg += f'，自动创建 {len(new_fields)} 个字段'
    return jsonify({'message': msg, 'count': ok, 'fail_count': fail, 'new_fields': new_fields})

@app.route('/api/import/preview', methods=['POST'])
def api_import_preview():
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    try: import openpyxl
    except ImportError: return jsonify({'error': '未安装 openpyxl'}), 500
    file = request.files.get('file')
    if not file: return jsonify({'error': '请上传文件'}), 400
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    max_r = min(ws.max_row, 20)
    preview_rows = []
    for ri in range(1, max_r + 1):
        cells = [str(c.value).strip() if c.value is not None else '' for c in ws[ri]]
        while cells and cells[-1] == '': cells.pop()
        if any(c for c in cells if c): preview_rows.append({'row': ri, 'cells': cells, 'non_empty': sum(1 for c in cells if c)})
    best_row = _detect_real_header_row(ws, 1)
    return jsonify({'preview_rows': preview_rows, 'detected_header_row': best_row, 'suggested_header_row': best_row})

# ──────────────────────── 备份恢复 ────────────────────────
@app.route('/api/backup', methods=['GET'])
def api_list_backups():
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(session['user_id'], 'export_excel'): return jsonify({'error': '权限不足'}), 403
    rows = get_db().fetch_all("SELECT * FROM backups ORDER BY created_at DESC")
    return jsonify([{'filename': r['filename'], 'size_bytes': r['size_bytes'], 'created_at': r['created_at']} for r in rows])

@app.route('/api/backup', methods=['POST'])
def api_create_backup():
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(session['user_id'], 'export_excel'): return jsonify({'error': '权限不足'}), 403
    g.current_user_id = session['user_id']; g.current_username = session.get('username', ''); g.current_role = session.get('role', '')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'backup_{timestamp}.db'
    src_path = _get_real_db_path()
    if not os.path.exists(src_path): return jsonify({'error': '数据库文件不存在'}), 500
    # 复制数据库文件到备份目录
    dest_path = os.path.join(BACKUP_DIR, filename)
    shutil.copy2(src_path, dest_path)
    size = os.path.getsize(dest_path)
    db = get_db()
    db.execute_query("INSERT INTO backups (filename, size_bytes) VALUES (?,?)", (filename, size))
    log_operation('创建备份', target_type='backup', target_id=filename, detail=json.dumps({'size': size}, ensure_ascii=False))
    return jsonify({'message': '备份创建成功', 'filename': filename, 'size_bytes': size})

@app.route('/api/backup/<filename>/download', methods=['GET'])
def api_download_backup(filename):
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    file_path = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(file_path): return jsonify({'error': '备份文件不存在'}), 404
    return send_file(file_path, mimetype='application/octet-stream', as_attachment=True, download_name=filename)

@app.route('/api/backup/<filename>', methods=['DELETE'])
def api_delete_backup(filename):
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    db = get_db()
    db.execute_query("DELETE FROM backups WHERE filename = ?", (filename,))
    # 同时删除备份文件
    file_path = os.path.join(BACKUP_DIR, filename)
    if os.path.exists(file_path):
        os.remove(file_path)
    return jsonify({'message': '备份已删除'})

@app.route('/api/backup/restore', methods=['POST'])
@require_role('boss')
def api_restore_backup():
    file = request.files.get('file')
    if not file: return jsonify({'error': '请上传数据库文件'}), 400
    db_path = _get_real_db_path()
    if os.path.exists(db_path):
        shutil.copy2(db_path, db_path + f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db')
    reset_adapter()
    file.save(db_path)
    log_operation('恢复数据库备份')
    return jsonify({'message': '数据库恢复成功'})

# ──────────────────────── 用户管理 ────────────────────────
@app.route('/api/users', methods=['GET'])
def api_users_list():
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    if not user_has_permission(session['user_id'], 'manage_users'): return jsonify({'error': '权限不足'}), 403
    users = get_db().fetch_all("SELECT id, username, role, permissions, COALESCE(global_audit, 0) AS audit_enabled, created_at FROM users ORDER BY created_at ASC")
    result = []
    for u in users:
        try: perms = json.loads(u['permissions'] or '[]')
        except: perms = []
        result.append({'id': u['id'], 'username': u['username'], 'role': u['role'], 'permissions': perms, 'audit_enabled': bool(u['audit_enabled']), 'created_at': u['created_at']})
    return jsonify(result)

@app.route('/api/users', methods=['POST'])
@require_perm('manage_users')
def api_create_user():
    data = request.get_json(force=True) or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    user_role = data.get('role', 'employee')
    permissions = data.get('permissions')
    audit_enabled = data.get('audit_enabled', False)
    if not username or len(password) < 6: return jsonify({'error': '用户名不能为空，密码至少6位'}), 400
    if user_role not in ('boss', 'hr', 'employee', 'custom'): return jsonify({'error': f'角色无效: {user_role}'}), 400
    if user_role == 'custom': user_role = 'custom'
    user_id = session.get('user_id')
    session_role = session.get('role', '')
    db = get_db()
    if db.fetch_one("SELECT id FROM users WHERE username = ?", (username,)): return jsonify({'error': '用户名已存在'}), 400
    if session_role != 'boss' and is_user_audited(user_id):
        db.execute_query("INSERT INTO pending_changes (row_id, column_name, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?)",
                         (None, username, json.dumps(data, ensure_ascii=False), 'add_user', user_id, '待审核'))
        new_id = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
        log_operation('提交添加用户审核', target_type='pending_changes', target_id=str(new_id))
        return jsonify({'message': '添加用户申请已提交审核', 'pending': True}), 200
    if permissions is not None and isinstance(permissions, list):
        perms = json.dumps(permissions, ensure_ascii=False)
    elif user_role == 'custom':
        perms = json.dumps(permissions or [], ensure_ascii=False)
    else:
        perms = json.dumps(ROLE_DEFAULT_PERMISSIONS.get(user_role, []), ensure_ascii=False)
    global_audit = 1 if audit_enabled else 0
    db.execute_query("INSERT INTO users (username, password_hash, role, permissions, global_audit) VALUES (?,?,?,?,?)",
                     (username, hash_password(password), user_role, perms, global_audit))
    new_uid = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
    log_operation('添加用户', target_type='users', target_id=str(new_uid), detail=json.dumps({'username': username, 'role': user_role}, ensure_ascii=False))
    return jsonify({'message': '用户创建成功', 'id': new_uid}), 201

@app.route('/api/users/<int:uid>', methods=['PUT'])
@require_perm('manage_users')
def api_update_user(uid):
    data = request.get_json(force=True) or {}
    username = (data.get('username') or '').strip()
    role = data.get('role', 'employee')
    permissions = data.get('permissions')
    audit_enabled = data.get('audit_enabled')
    if not username: return jsonify({'error': '用户名不能为空'}), 400
    # 保持 custom 角色不转换
    user_id = session.get('user_id')
    session_role = session.get('role', '')
    db = get_db()
    if session_role != 'boss' and is_user_audited(user_id):
        audit_data = {'id': uid, 'username': username, 'role': role, 'password': data.get('password'), 'permissions': permissions, 'audit_enabled': audit_enabled}
        db.execute_query("INSERT INTO pending_changes (row_id, column_name, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?)",
                         (uid, username, json.dumps(audit_data, ensure_ascii=False), 'update_user', user_id, '待审核'))
        new_id = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
        log_operation('提交修改用户审核', target_type='pending_changes', target_id=str(new_id))
        return jsonify({'message': '修改用户申请已提交审核', 'pending': True}), 200
    # 构建 UPDATE 语句
    set_parts = []
    params = []
    set_parts.append("username = ?"); params.append(username)
    set_parts.append("role = ?"); params.append(role)
    if data.get('password'):
        set_parts.append("password_hash = ?"); params.append(hash_password(data['password']))
    if permissions is not None and isinstance(permissions, list):
        set_parts.append("permissions = ?"); params.append(json.dumps(permissions, ensure_ascii=False))
    if audit_enabled is not None:
        set_parts.append("global_audit = ?"); params.append(1 if audit_enabled else 0)
    params.append(uid)
    db.execute_query(f"UPDATE users SET {', '.join(set_parts)} WHERE id = ?", tuple(params))
    log_operation('修改用户', target_type='users', target_id=str(uid))
    return jsonify({'message': '用户更新成功'})

@app.route('/api/users/<int:uid>', methods=['DELETE'])
@require_perm('manage_users')
def api_delete_user(uid):
    user_id = session.get('user_id')
    session_role = session.get('role', '')
    db = get_db()
    u = db.fetch_one("SELECT username FROM users WHERE id=? AND role!='boss'", (uid,))
    if not u: return jsonify({'error': '用户不存在或无法删除'}), 404
    if session_role != 'boss' and is_user_audited(user_id):
        db.execute_query("INSERT INTO pending_changes (row_id, column_name, new_value, change_type, requested_by, status) VALUES (?,?,?,?,?,?)",
                         (uid, u['username'], json.dumps({'id': uid, 'username': u['username']}, ensure_ascii=False), 'delete_user', user_id, '待审核'))
        new_id = db.fetch_one("SELECT last_insert_rowid() AS id")['id']
        log_operation('提交删除用户审核', target_type='pending_changes', target_id=str(new_id))
        return jsonify({'message': '删除用户申请已提交审核', 'pending': True}), 200
    db.execute_query("DELETE FROM users WHERE id=?", (uid,))
    log_operation('删除用户', target_type='users', target_id=str(uid))
    return jsonify({'message': f'用户「{u["username"]}」已删除'})

# ──────────────────────── 设置 ────────────────────────
def _get_config_value(db, key: str, default: str = '') -> str:
    """从 system_config 表读取配置值"""
    row = db.fetch_one("SELECT `value` FROM system_config WHERE `key` = ?", (key,))
    return row['value'] if row else default

def _set_config_value(db, key: str, value: str):
    """向 system_config 表写入或更新配置值（先删后插，兼容 SQLite / MySQL）"""
    db.execute_query("DELETE FROM system_config WHERE `key` = ?", (key,))
    db.execute_query("INSERT INTO system_config (`key`, `value`) VALUES (?, ?)", (key, value))

@app.route('/api/settings', methods=['GET'])
def api_settings():
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    db = get_db()
    engine = _get_config_value(db, 'db_engine', 'sqlite')
    mysql_host = _get_config_value(db, 'mysql_host', 'localhost')
    mysql_port = int(_get_config_value(db, 'mysql_port', '3306'))
    mysql_user = _get_config_value(db, 'mysql_user', 'root')
    mysql_password = _get_config_value(db, 'mysql_password', '')
    mysql_database = _get_config_value(db, 'mysql_database', 'app_db')
    return jsonify({
        'db_engine': engine,
        'mysql_host': mysql_host,
        'mysql_port': mysql_port,
        'mysql_user': mysql_user,
        'mysql_password': mysql_password,
        'mysql_database': mysql_database,
    })

@app.route('/api/settings', methods=['PUT'])
def api_update_settings():
    if not session.get('user_id'): return jsonify({'error': '请先登录'}), 401
    db = get_db()
    data = request.get_json(force=True) or {}
    if 'db_engine' in data:
        _set_config_value(db, 'db_engine', data['db_engine'])
    if 'mysql_host' in data:
        _set_config_value(db, 'mysql_host', data['mysql_host'])
    if 'mysql_port' in data:
        _set_config_value(db, 'mysql_port', str(data['mysql_port']))
    if 'mysql_user' in data:
        _set_config_value(db, 'mysql_user', data['mysql_user'])
    if 'mysql_password' in data:
        _set_config_value(db, 'mysql_password', data['mysql_password'])
    if 'mysql_database' in data:
        _set_config_value(db, 'mysql_database', data['mysql_database'])
    log_operation('更新系统设置', target_type='system_config', detail=json.dumps(data, ensure_ascii=False))
    return jsonify({'message': '设置已保存，切换数据库引擎需重启服务后生效'})

@app.route('/api/settings/test-mysql', methods=['POST'])
def test_mysql_connection():
    data = request.get_json(force=True) or {}
    host = (data.get('host') or '').strip()
    port = data.get('port')
    user = (data.get('user') or '').strip()
    password = data.get('password', '')
    database = (data.get('database') or '').strip()

    # 参数校验
    if not host:
        return jsonify({'success': False, 'message': '缺少参数: host（主机地址）'}), 400
    if not port:
        return jsonify({'success': False, 'message': '缺少参数: port（端口号）'}), 400
    if not user:
        return jsonify({'success': False, 'message': '缺少参数: user（用户名）'}), 400
    if not database:
        return jsonify({'success': False, 'message': '缺少参数: database（数据库名）'}), 400

    try:
        port = int(port)
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': f'端口号格式错误: {port}'}), 400

    print(f'[测试连接] host={host}, port={port}, user={user}, password={"***" if password else "(空)"}, database={database}')

    from backend.db_adapter import test_mysql_connection as do_test
    result = do_test(host=host, port=port, user=user, password=password, database=database)
    if result['success']:
        print(f'[测试连接] MySQL 连接成功: {host}:{port}/{database}')
        return jsonify(result)
    else:
        print(f'[测试连接] MySQL 连接失败: {result["message"]}')
        return jsonify(result), 200

# ──────────────────────── AI 数据分析 API ────────────────────────

@app.route('/api/ai/models', methods=['GET'])
@app.route('/api/ai/models/', methods=['GET'])
def get_ai_models():
    if not session.get('user_id'):
        return jsonify({'error': '请先登录'}), 401
    """获取支持的 AI 模型预设列表"""
    from backend.ai_client import MODEL_PRESETS
    # 返回预设信息（不包含敏感数据）
    presets = {}
    for key, preset in MODEL_PRESETS.items():
        presets[key] = {
            "name": preset["name"],
            "models": preset["models"],
            "base_url": preset["base_url"],
            "description": preset["description"],
        }
    return jsonify({"success": True, "models": presets})


def _safe_close(adapter):
    """安全关闭数据库适配器，忽略异常（可能未初始化或已关闭）"""
    try:
        if adapter is not None:
            adapter.close()
    except Exception:
        pass


def _attach_warning(payload: dict, warning):
    """若 warning 非空，将其写入响应字典顶层（前端可读取 _warning 字段做提示）"""
    if warning:
        payload['_warning'] = warning
    return payload


def _parse_ai_request(data: dict, default_model: str = 'gpt-4o-mini'):
    """
    从请求体解析 AI 调用参数并创建 AIClient（供 test / analyze / chat 三个接口复用）。

    Args:
        data: 请求体 JSON 字典
        default_model: 前端未指定 model 时使用的默认模型名

    Returns:
        tuple (provider, model, client, error_response, warning)：
        - 校验失败时：client 为 None，error_response 为 (jsonify响应, HTTP状态码) 元组
        - 校验成功时：error_response 为 None，client 为已配置好的 AIClient 实例
        - warning：可选字符串，提示前端展示的兼容性提示（如 web_search 不被支持时）
    """
    from backend.ai_client import AIClient, MODEL_PRESETS

    provider = data.get('provider', 'openai')
    model = data.get('model', default_model)
    api_key = data.get('api_key', '')
    base_url = data.get('base_url', '')
    temperature = data.get('temperature', 0.7)
    web_search = bool(data.get('web_search', False))
    warning = None

    # 联网搜索仅 DeepSeek 官方 chat/completions 端点支持 tools[0].type='web_search'，
    # 其他 provider（OpenAI / Qwen / ERNIE / 自定义中转）走 /chat/completions 协议，
    # 不接受该变体，强制开启会被服务端以 400 拒绝（unknown variant 'web_search'）。
    # 判断依据：provider 必须是 deepseek，且 Base URL 必须指向 DeepSeek 官方域名。
    effective_base_url = base_url or (MODEL_PRESETS.get(provider, {}).get("base_url", ""))
    is_deepseek_official = provider == 'deepseek' and 'deepseek.com' in effective_base_url.lower()
    if web_search and not is_deepseek_official:
        web_search = False
        warning = '联网搜索仅 DeepSeek 官方端点支持，已自动关闭'

    # 无效输入校验：缺少 API Key 时直接返回友好错误，避免无意义的网络请求
    if not api_key:
        return None, None, None, (jsonify({'success': False, 'error': '请提供 API Key'}), 400), None

    client = AIClient(
        provider=provider,
        model=model,
        api_key=api_key,
        base_url=base_url if base_url else None,
        temperature=temperature,
        web_search=web_search,
    )
    return provider, model, client, None, warning


@app.route('/api/ai/test', methods=['POST'])
@app.route('/api/ai/test/', methods=['POST'])
@require_login
def test_ai_connection():
    """测试 AI 连接"""
    data = request.get_json() or {}
    provider, model, client, err_resp, warning = _parse_ai_request(data, 'gpt-4o-mini')
    if err_resp:
        return err_resp

    try:
        result = client.test_connection()

        # 记录操作日志
        if result.get('success'):
            log_operation('AI连接测试', 'ai_test', detail=f'模型: {model}, 厂商: {provider}')
        else:
            log_operation('AI连接测试失败', 'ai_test', detail=f'模型: {model}, 厂商: {provider}, 错误: {result.get("error", "未知错误")[:100]}')

        return jsonify(_attach_warning(result, warning))
    except Exception as e:
        log_operation('AI连接测试异常', 'ai_test', detail=f'异常: {str(e)[:200]}')
        return jsonify({'success': False, 'error': f'测试失败: {str(e)}'}), 200


@app.route('/api/ai/analyze', methods=['POST'])
@app.route('/api/ai/analyze/', methods=['POST'])
@require_login
def ai_analyze():
    """AI 数据分析"""
    data = request.get_json() or {}
    provider, model, client, err_resp, warning = _parse_ai_request(data, 'gpt-4.1-mini')
    if err_resp:
        return err_resp
    question = data.get('question', '')
    adapter = None

    try:
        adapter = get_adapter()

        # ── 1. 列定义（columns_meta 表）──
        columns = []
        try:
            columns = adapter.fetch_all("SELECT * FROM columns_meta ORDER BY position ASC")
            columns = [dict(c) for c in columns] if columns else []
        except Exception:
            pass

        # ── 2. 总行数（rows_data 表，按角色权限过滤）──
        # 管理角色（boss/hr）统计全部数据；员工只能统计自己创建的数据
        current_role = g.current_role or session.get('role', '')
        current_username = g.current_username or session.get('username', '')
        total_rows = 0
        try:
            if can_view_all_rows(current_role):
                result = adapter.fetch_one("SELECT COUNT(*) as cnt FROM rows_data")
            else:
                result = adapter.fetch_one("SELECT COUNT(*) as cnt FROM rows_data WHERE _created_by = ?", (current_username,))
            total_rows = result['cnt'] if result else 0
        except Exception:
            total_rows = 0

        # ── 3. 字段统计 ──
        field_stats = []
        if columns:
            for col in columns:
                col_name = col.get('name', '')
                col_label = col.get('label', col_name)
                col_type = col.get('field_type', 'text')
                if not col_name:
                    continue
                try:
                    if can_view_all_rows(current_role):
                        rows = adapter.fetch_all(
                            f"SELECT {col_name} as val FROM rows_data LIMIT 1000"
                        )
                    else:
                        rows = adapter.fetch_all(
                            f"SELECT {col_name} as val FROM rows_data WHERE _created_by = ? LIMIT 1000",
                            (current_username,),
                        )
                    if rows:
                        values = [r['val'] for r in rows if r.get('val') is not None]
                        stats = {
                            "总记录数": len(rows),
                            "有值记录数": len(values),
                            "空值率": f"{round((1 - len(values) / len(rows)) * 100, 1)}%" if rows else "0%",
                        }
                        if col_type in ('number',):
                            nums = [float(v) for v in values if v is not None]
                            if nums:
                                stats["最小值"] = min(nums)
                                stats["最大值"] = max(nums)
                                stats["平均值"] = round(sum(nums) / len(nums), 2)
                        elif col_type in ('select',):
                            counter = Counter([str(v) for v in values])
                            stats["分布"] = dict(counter.most_common(10))
                        field_stats.append({
                            "key": col_name,
                            "label": col_label,
                            "type": col_type,
                            "stats": stats,
                        })
                except Exception:
                    # 字段可能不存在或查询失败，跳过
                    pass

        # ── 4. 审核统计 ──
        audit_stats = {}
        try:
            pending = adapter.fetch_one("SELECT COUNT(*) as cnt FROM pending_changes WHERE status='待审核'")
            approved = adapter.fetch_one("SELECT COUNT(*) as cnt FROM pending_changes WHERE status='已通过'")
            rejected = adapter.fetch_one("SELECT COUNT(*) as cnt FROM pending_changes WHERE status='已驳回'")
            audit_stats = {
                "待审核": pending['cnt'] if pending else 0,
                "已通过": approved['cnt'] if approved else 0,
                "已驳回": rejected['cnt'] if rejected else 0,
            }
        except Exception:
            pass

        # ── 5. 最近操作日志 ──
        recent_logs = []
        try:
            logs = adapter.fetch_all(
                "SELECT * FROM operations_log ORDER BY created_at DESC LIMIT 20"
            )
            recent_logs = [dict(log) for log in (logs or [])]
        except Exception:
            pass

        # ── 6. 用户统计 ──
        user_stats = {}
        try:
            total_users = adapter.fetch_one("SELECT COUNT(*) as cnt FROM users")
            boss_count = adapter.fetch_one("SELECT COUNT(*) as cnt FROM users WHERE role='boss'")
            hr_count = adapter.fetch_one("SELECT COUNT(*) as cnt FROM users WHERE role='hr'")
            employee_count = adapter.fetch_one("SELECT COUNT(*) as cnt FROM users WHERE role='employee'")
            user_stats = {
                "总用户数": total_users['cnt'] if total_users else 0,
                "管理员": boss_count['cnt'] if boss_count else 0,
                "HR": hr_count['cnt'] if hr_count else 0,
                "员工": employee_count['cnt'] if employee_count else 0,
            }
        except Exception:
            pass

        # ── 构造数据摘要（按当前用户权限裁剪敏感部分）──
        # 员工 / 无对应权限的用户，不得把审核统计、用户统计、操作日志等
        # 管理类数据喂给 AI，避免“套出越权信息”的问题。
        current_perms = get_user_permissions(g.current_user_id) if g.current_user_id else []
        def _has_perm(perm: str) -> bool:
            return perm in current_perms

        data_summary = {
            "total_rows": total_rows,
            "total_columns": len(columns),
            "columns": [{"key": c.get('name', ''), "label": c.get('label', c.get('name', '')), "type": c.get('field_type', 'text')} for c in columns],
            "field_stats": field_stats,
            "audit_stats": audit_stats if _has_perm('audit_center') else {},
            "user_stats": user_stats if _has_perm('manage_users') else {},
            "recent_logs": recent_logs if _has_perm('view_logs') else [],
        }

        # 用户身份与权限上下文：让 AI 明确“提问者是谁、能看到什么、分析范围多大”
        user_context = {
            "username": g.current_username,
            "role": current_role,
            "permissions": current_perms,
            "data_scope": "all" if can_view_all_rows(current_role) else "own",
        }

        # ── 调用 AI（client 已在 _parse_ai_request 中创建）──
        result = client.analyze(data_summary, question=question if question else None, user_context=user_context)

        # 记录操作日志
        if result.get('success'):
            log_operation(
                action='AI分析',
                target_type='ai_analyze',
                detail=f'模型: {model}, 厂商: {provider}' + (f', 问题: {question[:80]}' if question else '')
            )
        else:
            log_operation(
                action='AI分析失败',
                target_type='ai_analyze',
                detail=f'模型: {model}, 厂商: {provider}, 错误: {result.get("error", "未知错误")[:100]}'
            )

        return jsonify(_attach_warning(result, warning))

    except Exception as e:
        traceback.print_exc()
        log_operation('AI分析异常', 'ai_analyze', detail=f'异常: {str(e)[:200]}')
        return jsonify({'success': False, 'error': f'分析失败: {str(e)}'}), 500
    finally:
        _safe_close(adapter)


@app.route('/api/ai/chat', methods=['POST'])
@app.route('/api/ai/chat/', methods=['POST'])
@require_login
def ai_chat():
    """AI 对话式追问"""
    data = request.get_json() or {}
    provider, model, client, err_resp, warning = _parse_ai_request(data, 'gpt-4o-mini')
    if err_resp:
        return err_resp
    history = data.get('history', [])
    message = data.get('message', '')

    if not message:
        return jsonify({'success': False, 'error': '请输入消息'}), 400

    try:
        result = client.chat(history, message)

        # 记录操作日志
        if result.get('success'):
            log_operation(
                action='AI对话',
                target_type='ai_chat',
                detail=f'模型: {model}, 厂商: {provider}, 消息: {message[:80]}'
            )
        else:
            log_operation(
                action='AI对话失败',
                target_type='ai_chat',
                detail=f'模型: {model}, 厂商: {provider}, 错误: {result.get("error", "未知错误")[:100]}'
            )

        return jsonify(_attach_warning(result, warning))
    except Exception as e:
        log_operation('AI对话异常', 'ai_chat', detail=f'异常: {str(e)[:200]}')
        return jsonify({'success': False, 'error': f'对话失败: {str(e)}'}), 500


# ──────────────────────── AI 分析记录（按用户隔离） ────────────────────────

def _get_ai_history_row(adapter, user_id):
    """获取当前用户最近一条 AI 分析记录（每个账号独立）"""
    return adapter.fetch_one(
        "SELECT report, chat_history, created_at, updated_at "
        "FROM ai_analysis_history WHERE user_id = ? "
        "ORDER BY id DESC LIMIT 1",
        (user_id,)
    )


@app.route('/api/ai/history', methods=['GET'])
@app.route('/api/ai/history/', methods=['GET'])
@require_login
def get_ai_history():
    """获取当前用户的 AI 分析记录（按用户隔离）"""
    try:
        adapter = get_db()
        row = _get_ai_history_row(adapter, g.current_user_id)
        if not row:
            return jsonify({'success': True, 'report': '', 'chat_history': [], 'created_at': None, 'updated_at': None})
        report = row['report'] or ''
        chat_history = []
        if row['chat_history']:
            try:
                chat_history = json.loads(row['chat_history'])
            except Exception:
                chat_history = []
        return jsonify({
            'success': True,
            'report': report,
            'chat_history': chat_history,
            'created_at': row['created_at'],
            'updated_at': row['updated_at'],
        })
    except Exception as e:
        log_operation('AI记录获取异常', 'ai_history', detail=f'异常: {str(e)[:200]}')
        return jsonify({'success': False, 'error': f'获取记录失败: {str(e)}'}), 500


@app.route('/api/ai/history', methods=['POST'])
@app.route('/api/ai/history/', methods=['POST'])
@require_login
def save_ai_history():
    """保存当前用户的 AI 分析记录（按用户隔离，只保留一份最新记录）"""
    data = request.get_json() or {}
    report = data.get('report', '')
    chat_history = data.get('chat_history', [])

    try:
        adapter = get_db()
        existing = _get_ai_history_row(adapter, g.current_user_id)
        if existing:
            adapter.execute_query(
                "UPDATE ai_analysis_history SET report = ?, chat_history = ?, updated_at = ({ts}) "
                "WHERE user_id = ?",
                (report, json.dumps(chat_history, ensure_ascii=False), g.current_user_id)
            )
        else:
            adapter.execute_query(
                "INSERT INTO ai_analysis_history (user_id, report, chat_history) VALUES (?, ?, ?)",
                (g.current_user_id, report, json.dumps(chat_history, ensure_ascii=False))
            )
        log_operation('AI记录保存', 'ai_history', detail=f'保存 {len(report)} 字报告 / {len(chat_history)} 条对话')
        return jsonify({'success': True})
    except Exception as e:
        log_operation('AI记录保存异常', 'ai_history', detail=f'异常: {str(e)[:200]}')
        return jsonify({'success': False, 'error': f'保存记录失败: {str(e)}'}), 500


@app.route('/api/ai/history', methods=['DELETE'])
@app.route('/api/ai/history/', methods=['DELETE'])
@require_login
def clear_ai_history():
    """清除当前用户的 AI 分析记录（按用户隔离）"""
    try:
        adapter = get_db()
        adapter.execute_query(
            "DELETE FROM ai_analysis_history WHERE user_id = ?",
            (g.current_user_id,)
        )
        log_operation('AI记录清除', 'ai_history', detail='清除该用户全部 AI 记录')
        return jsonify({'success': True})
    except Exception as e:
        log_operation('AI记录清除异常', 'ai_history', detail=f'异常: {str(e)[:200]}')
        return jsonify({'success': False, 'error': f'清除记录失败: {str(e)}'}), 500


# 前端构建产物路径（新 UI 使用 Vite 构建到 frontend/dist）
FRONTEND_DIST = os.path.join(base_path, 'frontend', 'dist')

def _get_index_html():
    """每次请求时重新读取 index.html，确保最新构建的 JS/CSS 引用生效"""
    index_path = os.path.join(FRONTEND_DIST, 'index.html')
    if not os.path.exists(index_path):
        return None
    with open(index_path, 'r', encoding='utf-8') as _f:
        raw = _f.read()
    return raw


# ──────────────────────── React SPA 路由 ────────────────────────
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_react(path):
    if path.startswith('api/'):
        return jsonify({'error': 'Not found'}), 404
    if path.startswith('static/'):
        # 仍从 static/ 目录提供上传文件等静态资源
        response = send_from_directory('static', path[len('static/'):])
        response.headers['Cache-Control'] = 'public, max-age=3600'
        return response
    if path.startswith('assets/') or '.' in path.split('/')[-1]:
        # 前端资源（JS/CSS）从 Vite 构建目录提供
        file_path = os.path.join(FRONTEND_DIST, path)
        if os.path.exists(file_path):
            response = send_from_directory(FRONTEND_DIST, path)
            # 使用 ETag + max-age=0 代替 immutable，确保每次加载都验证版本
            response.headers['Cache-Control'] = 'public, max-age=0, must-revalidate'
            return response
        # 不加缓存破坏参数，避免找不到真实文件时造成额外问题
        response = send_from_directory(FRONTEND_DIST, 'index.html')
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Vary'] = '*'
        return response
    # 每次请求动态读取 index.html，确保最新构建生效
    _index_html = _get_index_html()
    if _index_html is not None:
        response = Response(_index_html, mimetype='text/html')
    else:
        response = send_from_directory(FRONTEND_DIST, 'index.html')
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    response.headers['Vary'] = '*'
    return response


# ──────────────────────── 启动 ────────────────────────
if __name__ == '__main__':
    try:
        import waitress
        port = 5001
        print(f'[INFO] 动态数据登记系统 - 生产模式启动 (Waitress)')
        print(f'   后端地址: http://0.0.0.0:{port}')
        print(f'   前端地址: http://localhost:5173')
        print(f'   线程数: 20 (支持 150+ 并发)')
        waitress.serve(app, host='0.0.0.0', port=port, threads=20, channel_timeout=120)
    except ImportError:
        print('[WARN] waitress 未安装，使用开发服务器。请执行: pip install waitress')
        port = 5001
        print(f'[INFO] 开发模式启动: http://127.0.0.1:{port}')
        app.run(host='127.0.0.1', port=port, debug=False, threaded=True)
    except Exception:
        with open('flask_crash.log', 'w', encoding='utf-8') as f:
            f.write(f"=== Crash at {datetime.now()} ===\n")
            f.write(traceback.format_exc())
        raise
